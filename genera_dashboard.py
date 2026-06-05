"""
FPS Dashboard Generator
Legge l'Excel e genera docs/index.html
Eseguito automaticamente da GitHub Actions ad ogni push dell'Excel.
"""
import pandas as pd
from datetime import datetime as _dt
import json
import math
import base64
import sys
import os
from datetime import datetime

# ─── TROVA IL FILE EXCEL ────────────────────────────────────────────
def trova_excel():
    for f in os.listdir('.'):
        if f.endswith('.xlsx') and not f.startswith('~'):
            return f
    print("ERRORE: nessun file .xlsx trovato nella cartella!")
    sys.exit(1)

EXCEL = trova_excel()
print(f"Lettura: {EXCEL}")

# ─── HELPERS ────────────────────────────────────────────────────────
def n(v):
    try:
        x = float(str(v).replace(',', '.').replace(' ', ''))
        return 0.0 if (math.isnan(x) or math.isinf(x)) else x
    except:
        return 0.0

def ni(v):
    return int(n(v))

def s(v):
    return '' if v is None or str(v) == 'nan' else str(v).strip()

def isFB(v):
    v = s(v)
    if len(v) < 4 or not v[0].isupper():
        return False
    skip = ['Family', 'FBO', 'Gruppo', 'Appt', 'Polizze', 'Premio',
            'YTD', 'Periodo', 'Data', 'Celle', 'Modifica', 'TOT', 'NaN']
    return not any(v.startswith(x) for x in skip)

def fe(v):
    v = 0 if (v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v)))) else v
    r = round(v)
    # Format with dot as thousands separator: 1234 -> 1.234, 12345 -> 12.345
    s = f"{r:,}".replace(",", ".")
    return "\u20ac\u00a0" + s

def fe_k(v):
    """Formato migliaia con punto, es: € 2.265"""
    v = 0 if (v is None or (isinstance(v, float) and (math.isnan(v) or math.isinf(v)))) else v
    return "€\u00a0" + f"{round(v):,}".replace(",", ".")

def fn(v):
    return f"{int(v):,}".replace(",", ".")

def fp(v):
    return str(round(v * 100)) + "%"

def tag(cls, txt):
    return f'<span class="tag {cls}">{txt}</span>'

def badge(cls, txt):
    return f'<span class="bdg {cls}">{txt}</span>'

def dot(state):
    colors = {
        "green":   "#2E8B5F",
        "amber":   "#D97706",
        "red":     "#C0392B",
        "neutral": "#94A3B8"
    }
    c = colors.get(state, "#94A3B8")
    return f'<span style="color:{c};font-size:1.1rem">&#x25CF;</span>'

# ─── LEGGI EXCEL ────────────────────────────────────────────────────
xl = pd.read_excel(EXCEL, sheet_name=None)
D = {}

# KPI Generale
kg = xl['📊 KPI Generale']
r6, r10, r14, r15 = kg.iloc[6].to_list(), kg.iloc[10].to_list(), kg.iloc[14].to_list(), kg.iloc[15].to_list()
D['kpiGen'] = {
    'target': n(r6[3]), 'gap31dic': n(r6[6]),
    'polizze': ni(r10[1]), 'premioAnnuo': round(n(r10[2]), 2),
    'incassoPrevisto': round(n(r10[3]), 2), 'premioFirma': round(n(r10[4]), 2),
    'premiIncassati': round(n(r10[5]), 2), 'residuo2026': round(n(r10[6]), 2),
    'residuo2027': round(n(r10[7]), 2), 'totResiduo': round(n(r10[8]), 2),
    'premioFirmaYTD': round(n(r10[4]), 2),
    'polLav': ni(r14[1]), 'firmaLav': round(n(r14[2]), 2),
    'apptCom': ni(r14[3]), 'apptEff': ni(r14[4]),
    'convRate': round(n(r14[5]), 4), 'callback': ni(r14[6]),
    'provv': round(n(r14[7]), 2), 'fbAttivi': ni(r14[8])
}

# Parse fbTot from string like '42% su 19 FB'
import re as _re
_fbt = _re.search(r'su\s+(\d+)', str(D['kpiGen'].get('fbTotStr','')))
D['kpiGen']['fbTot'] = int(_fbt.group(1)) if _fbt else 19

# KPI Backup
kb_sheet = xl['💾 KPI Backup']
D['kpiBackup'] = {}
MESE_CORR = datetime.today().month
for i in range(7, 35):
    try:
        r = kb_sheet.iloc[i].to_list()
    except:
        break
    nm = s(r[2])
    if not isFB(nm):
        continue
    pol = ni(r[4])
    obj = ni(r[11])
    if obj == 0:
        stato = 'neutral'
    elif pol >= MESE_CORR:
        stato = 'green'
    elif pol >= MESE_CORR - 2:
        stato = 'amber'
    else:
        stato = 'red'
    D['kpiBackup'][nm] = {
        'fbo': s(r[0]), 'gruppo': s(r[1]),
        'apptTot': ni(r[3]), 'polizze': pol, 'nonSott': ni(r[5]),
        'callback': ni(r[6]), 'conv': round(n(r[7]), 4),
        'premioAnnuo': round(n(r[8]), 2), 'incassoResiduo': round(n(r[9]), 2),
        'provv': round(n(r[10]), 2), 'objMese': obj,
        'deltaObj': ni(round(n(r[12]))), 'premiIncassati': round(n(r[17]), 2),
        'statoAttivo': stato
    }

# Obj Onorato
oo = xl['🎯 Obj Onorato']
r2 = oo.iloc[2].to_list()
D['objOnorato'] = {
    'obiettivo': n(r2[1]), 'incassato': round(n(r2[2]), 2),
    'deltaOggi': round(n(r2[3]), 2), 'previstoDic': round(n(r2[4]), 2),
    'deltaDic': round(n(r2[5]), 2)
}

# Collaboratori
col_sheet = xl['👥 Collaboratori']
D['collaboratori'] = []
for i in range(3, 30):
    try:
        r = col_sheet.iloc[i].to_list()
    except:
        break
    nm = s(r[2])
    if not isFB(nm):
        continue
    ing = r[4].strftime('%d/%m/%Y') if hasattr(r[4], 'strftime') else s(r[4])
    D['collaboratori'].append({
        'fbo': s(r[0]), 'gruppo': s(r[1]), 'name': nm,
        'email': s(r[3]), 'ingresso': ing,
        'cellulare': s(r[10]) if len(r)>10 and r[10] else '',
        'objAppt': ni(r[5]), 'objMese': ni(r[6]), 'objPremio': ni(r[7])
    })

# Dati Giornalieri
gj_sheet = xl['📝 Dati Giornalieri']
D['giornalieri'] = []
for i in range(4, len(gj_sheet)):
    try:
        r = gj_sheet.iloc[i].to_list()
    except:
        break
    ta = s(r[4])
    fb = s(r[8])
    if not ta or not fb:
        continue
    dp = r[17].strftime('%d/%m/%Y') if hasattr(r[17], 'strftime') else s(r[17])
    pf = n(r[14])
    pa = n(r[16])
    D['giornalieri'].append({
        'mese': s(r[1]), 'fb': fb, 'cliente': s(r[10]),
        'esito': s(r[11]), 'callbackData': s(r[12]),
        'tipoPol': s(r[13]), 'premioFirma': 0.0 if math.isnan(pf) else round(pf, 2),
        'frazionamento': s(r[15]), 'premioAnnuo': 0.0 if math.isnan(pa) else round(pa, 2),
        'dataPolizza': dp, 'statoPolizza': s(r[20]), 'tipoAtt': ta,
        'data': r[0] if hasattr(r[0], 'strftime') else None
    })

# Controllo Rate
ct_sheet = xl['💳 Controllo Rate']
D['polLavorazione'] = []
for i in range(3, 400):
    try:
        r = ct_sheet.iloc[i].to_list()
    except:
        break
    if s(r[8]) == 'In lavorazione' and isFB(s(r[1])):
        D['polLavorazione'].append({
            'fb': s(r[1]), 'cliente': s(r[2]), 'tipoPol': s(r[3]),
            'premioFirma': round(n(r[4]), 2), 'premioAnnuo': round(n(r[7]), 2),
            'dataPolizza': r[6], 'incassoResiduo': round(n(r[24]), 2)
        })

# Challenge sheet
ch_sheet = xl['🏁 Challenge']
ch_rows = []
try:
    ch_inizio = ch_sheet.iloc[0].to_list()[6]  # Data inizio
    ch_fine   = ch_sheet.iloc[1].to_list()[6]  # Data fine
    ch_min_pol = s(ch_sheet.iloc[2].to_list()[6])
    ch_min_pa  = s(ch_sheet.iloc[3].to_list()[6])
    # Classifica
    for i in range(5, 30):
        r = ch_sheet.iloc[i].to_list()
        if r[0] and str(r[0]) not in ('nan','') and str(r[0]).strip().isdigit():
            ch_rows.append({'pos':s(r[0]),'name':s(r[1]),'pol':s(r[2]),'pa':s(r[3])})
    # Partecipanti
    ch_part = []
    for i in range(6, 30):
        r = ch_sheet.iloc[i].to_list()
        if r[6] and str(r[6]) not in ('nan',''):
            ch_part.append(s(r[6]))
    ch_inizio_str = ch_inizio.strftime('%d/%m/%Y') if hasattr(ch_inizio,'strftime') else s(ch_inizio)
    ch_fine_str   = ch_fine.strftime('%d/%m/%Y') if hasattr(ch_fine,'strftime') else s(ch_fine)
except Exception as e:
    ch_inizio_str = ch_fine_str = ch_min_pol = ch_min_pa = ''
    ch_rows = []; ch_part = []

# Ritmo Vendita
rv_sheet = xl['📈 Ritmo Vendita 2026']
D['ritmoFB'] = {}
for i in range(24, 55):
    try:
        r = rv_sheet.iloc[i].to_list()
    except:
        break
    nm = s(r[4])
    if not isFB(nm):
        continue
    D['ritmoFB'][nm] = {
        'apptEff': ni(r[5]), 'pol': ni(r[6]), 'conv': round(n(r[7]), 4),
        'premiInc': round(n(r[8]), 2), 'budget': round(n(r[9]), 2),
        'pctBudget': round(n(r[10]), 4), 'proiezione': round(n(r[14]), 2),
        'stato': s(r[17])
    }

# ─── CALCOLI DERIVATI ───────────────────────────────────────────────
MESI_ORD = ['Gennaio', 'Febbraio', 'Marzo', 'Aprile', 'Maggio', 'Giugno',
            'Luglio', 'Agosto', 'Settembre', 'Ottobre', 'Novembre', 'Dicembre']
mesi_raw = sorted(set(r['mese'] for r in D['giornalieri'] if r['mese']),
                  key=lambda m: next((i for i, x in enumerate(MESI_ORD) if m.startswith(x)), 99))

D['collaboratori'].sort(key=lambda c: c['name'])
FBS = [c['name'] for c in D['collaboratori']]
PM, AM, PZ = {}, {}, {}
for r in D['giornalieri']:
    fb, mese = r['fb'], r['mese']
    if not fb or not mese:
        continue
    if r['esito'] == 'Sottoscritto':
        PM.setdefault(fb, {})[mese] = PM.get(fb, {}).get(mese, 0) + 1
        PZ.setdefault(fb, []).append(r)
    if r['tipoAtt'] == 'Appuntamento':
        AM.setdefault(fb, {})[mese] = AM.get(fb, {}).get(mese, 0) + 1

MB = [m[:3].upper() for m in mesi_raw]
pol_mese = [sum((PM.get(fb, {}).get(m, 0)) for fb in FBS) for m in mesi_raw]

# Premio alla Firma per mese di sottoscrizione (data appuntamento)
pf_mese = [round(sum(r['premioFirma'] for r in D['giornalieri']
                     if r['mese'] == m and r['esito'] == 'Sottoscritto'), 2)
           for m in mesi_raw]

# Incassato per mese di decorrenza (data polizza nel Controllo Rate, solo Processata)
_inc_cr = {}
try:
    _ct_inc = xl['💳 Controllo Rate']
    import pandas as _pd2
    for _ii2 in range(3, 600):
        try: _rr2 = _ct_inc.iloc[_ii2].to_list()
        except: break
        _dp2 = _rr2[6]
        _sta2 = s(_rr2[8])
        if _sta2 != 'Processata': continue
        try:
            if hasattr(_dp2, 'to_pydatetime'): _dp2 = _dp2.to_pydatetime()
            if not isinstance(_dp2, __import__('datetime').datetime): continue
        except: continue
        _mn2 = MESI_ORD[_dp2.month - 1] + ' ' + str(_dp2.year)
        _inc_cr[_mn2] = _inc_cr.get(_mn2, 0) + n(_rr2[4])
except: pass

pi_mese = [round(_inc_cr.get(m, 0), 2) for m in mesi_raw]

# ─── SVG SMOOTH LINE CHART ──────────────────────────────────────────
def smooth_svg(mb, values, color):
    W, H = 620, 260
    PL, PR, PT, PB = 60, 20, 30, 44
    cw, ch = W - PL - PR, H - PT - PB
    n_pts = len(mb)
    maxV = max(values) if values else 1

    def xp(i): return PL + i * cw / max(n_pts - 1, 1)
    def yp(v): return PT + ch - (v / max(maxV, 1)) * ch

    def bezier(vals):
        pts = [(xp(i), yp(v)) for i, v in enumerate(vals)]
        d = f"M {pts[0][0]:.1f},{pts[0][1]:.1f}"
        for i in range(1, len(pts)):
            p0 = pts[max(i - 2, 0)]; p1 = pts[i - 1]
            p2 = pts[i]; p3 = pts[min(i + 1, len(pts) - 1)]
            cx1 = p1[0] + (p2[0] - p0[0]) / 5; cy1 = p1[1] + (p2[1] - p0[1]) / 5
            cx2 = p2[0] - (p3[0] - p1[0]) / 5; cy2 = p2[1] - (p3[1] - p1[1]) / 5
            d += f" C {cx1:.1f},{cy1:.1f} {cx2:.1f},{cy2:.1f} {p2[0]:.1f},{p2[1]:.1f}"
        return d, pts

    base_y = PT + ch
    svg = f'<svg viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" width="100%">'

    for i in range(5):
        v = maxV * i / 4
        y = yp(v)
        lbl = f'€{round(v / 1000)}k' if v >= 1000 else f'€{round(v)}'
        svg += f'<line x1="{PL}" x2="{PL+cw}" y1="{y:.1f}" y2="{y:.1f}" stroke="rgba(11,30,61,.06)" stroke-width="1"/>'
        svg += f'<text x="{PL-7}" y="{y+4:.1f}" text-anchor="end" font-size="10" fill="#94A3B8" font-family="Outfit,sans-serif">{lbl}</text>'

    for i, lbl in enumerate(mb):
        svg += f'<text x="{xp(i):.1f}" y="{PT+ch+22}" text-anchor="middle" font-size="12" fill="#64748B" font-family="Outfit,sans-serif" font-weight="600">{lbl}</text>'

    path_d, pts = bezier(values)
    area_d = path_d + f" L {pts[-1][0]:.1f},{base_y} L {pts[0][0]:.1f},{base_y} Z"
    svg += (f'<defs><linearGradient id="gA" x1="0" y1="0" x2="0" y2="1">'
            f'<stop offset="0%" stop-color="{color}" stop-opacity="0.2"/>'
            f'<stop offset="100%" stop-color="{color}" stop-opacity="0"/>'
            f'</linearGradient></defs>')
    svg += f'<path d="{area_d}" fill="url(#gA)"/>'
    svg += f'<path d="{path_d}" fill="none" stroke="{color}" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>'

    for i, (x, y) in enumerate(pts):
        v = values[i]
        lbl = f'€{round(v / 1000, 1)}k' if v >= 1000 else f'€{round(v)}'
        off_y = -10 if y > PT + 20 else 16
        svg += f'<circle cx="{x:.1f}" cy="{y:.1f}" r="5" fill="white" stroke="{color}" stroke-width="2.5"/>'
        svg += f'<text x="{x:.1f}" y="{y+off_y:.1f}" text-anchor="middle" font-size="11" fill="{color}" font-family="Outfit,sans-serif" font-weight="700">{lbl}</text>'

    svg += '</svg>'
    return svg


# Monthly incassato from Controllo Rate (cols 9-20 = Gen-Dic)
MESI_COLS = {'GEN':9,'FEB':10,'MAR':11,'APR':12,'MAG':13,'GIU':14,'LUG':15,'AGO':16,'SET':17,'OTT':18,'NOV':19,'DIC':20}
ct_sheet2 = xl['💳 Controllo Rate']
inc_mese = []
for m in MB:
    col_idx = MESI_COLS.get(m, None)
    if col_idx:
        total = 0
        for i in range(3, 300):
            try:
                r = ct_sheet2.iloc[i].to_list()
                if r[1] and str(r[1]).strip() not in ('nan','') and r[1] != '#':
                    v = n(r[col_idx])
                    total += v if not math.isnan(v) else 0
            except: break
        inc_mese.append(round(total, 2))
    else:
        inc_mese.append(0)

SVG_PA = smooth_svg(MB, pf_mese, "#0B1E3D")
SVG_PI = smooth_svg(MB, pi_mese, "#2E8B5F")

# ─── CARD HELPER ────────────────────────────────────────────────────
def card(lbl, val, sub="", cls="", ico="", prog=None, bdg=""):
    prog_html = ""
    if prog:
        w = min(max(prog[0], 0), 100)
        prog_html = f'<div class="prog"><div class="pb {prog[1]}" style="width:{w}%"></div></div>'
    return (f'<div class="card {cls}">'
            f'<div class="cico">{ico}</div>'
            f'<p class="cl">{lbl}</p>'
            f'<p class="cv">{val}</p>'
            f'<p class="csub">{sub}</p>'
            f'{prog_html}{bdg}</div>')


# ─── BUILD ALL HTML SECTIONS ─────────────────────────────────────────
G = D['kpiGen']
ON = D['objOnorato']
KB = D['kpiBackup']
RL = D['ritmoFB']
PL_data = D['polLavorazione']

pct_pi = round(G['premiIncassati'] / max(G['premioAnnuo'], 1) * 100)
pol_lav_pa  = sum(p['premioAnnuo'] for p in PL_data)
pol_lav_pf  = sum(p['premioFirma'] for p in PL_data)
pol_lav_2026 = sum(p['incassoResiduo'] for p in PL_data)
fbA = sum(1 for k in KB.values() if k['statoAttivo'] == 'green')
fbTot = 19
pct_fba = round(fbA / max(fbTot, 1) * 100)
pct_target = round(G['premiIncassati'] / max(G['target'], 1) * 100)
pctO = round(ON['incassato'] / max(ON['obiettivo'], 1) * 100)
pctP = round(ON['previstoDic'] / max(ON['obiettivo'], 1) * 100)

# KPI Row 1: Polizze Sottoscritte, Premio Annuo
kpi1 = (
    card("Polizze Sottoscritte YTD", fn(G['polizze']),
         f"{len(PL_data)} in lavorazione", "gold", "📋",
         bdg=badge("bb", "Anno in corso"))
    + card("Premio Annuo Totale", fe(G['premioAnnuo']),
           f"Medio: {fe(G['premioAnnuo'] // max(G['polizze'], 1))}", "", "💶")
)

# KPI Row 2: Premio Firma, Incasso Previsto, Premi Incassati, Residuo, Pol. Lavorazione
kpi2 = (
    card("Premio alla Firma YTD", fe(G['premioFirmaYTD']),
         "Premio totale alla sottoscrizione", "gold", "📑",
         bdg=badge("bn", "YTD"))
    + card("Incasso Previsto 2026", fe(G['incassoPrevisto']),
           "Incassi certi entro dicembre", "navy", "📈",
           bdg=badge("bn", "Previsto"))
    + card("Premi Incassati", fe(G['premiIncassati']),
           f"{pct_pi}% del premio annuo", "green", "✅",
           prog=(pct_pi, "pg"))
    + card("Residuo da Incassare 2026", fe(G['residuo2026']),
           f"+ {fe(G['residuo2027'])} nel 2027", "red", "📉",
           bdg=badge("ba", f"Tot: {fe(G['totResiduo'])}"))
    + card("Polizze in Lavorazione", f"{fn(len(PL_data))} <span style='font-size:.65rem;background:var(--n3);color:#fff;padding:2px 8px;border-radius:10px;vertical-align:middle'>Da processare</span>",
           f"Premio Annuo: {fe(pol_lav_pa)}<br><span style='font-size:.75rem;color:var(--mut)'>Premio Firma tot.: {fe(pol_lav_pf)}</span><br><span style='font-size:.75rem;color:var(--mut)'>Incasso previsto 2026: {fe(pol_lav_2026)}</span>", "", "⚙️")
)

# KPI Row 3: Appt Comunicati, Tasso Conv, FB Attivi, Callback
kpi3 = (
    card("Appuntamenti Comunicati", fn(G['apptCom']),
         f"Effettuati: {fn(G['apptEff'])} ({round(G['apptEff']/max(G['apptCom'],1)*100)}%)", "", "📅")
    + card("Tasso di Conversione", fp(G['convRate']),
           f"{fn(G['apptEff'])} appt → {fn(G['polizze'])} polizze", "gold", "🎯",
           bdg=badge("bg", "Eccellente"))
    + card("Family Banker Attivi",
           f'{fbA}<span style="font-size:1rem;color:#64748B">/{fbTot}</span>',
           "Verde≥4pol · Arancione≥2pol · Rosso<2pol", "", "👥",
           prog=(pct_fba, "pa"))
    + card("Callback Aperti", fn(G['callback']),
           "Opportunità da richiamare subito", "amb", "🔄",
           bdg=badge("ba", "Urgente"))
)

# Onorato
onb = (
    f'<div><p class="ohl">Obiettivo Annuo</p><p class="ohv" style="color:#0B1E3D">{fe(ON["obiettivo"])}</p><p class="ohs">Budget 2026</p></div>'
    f'<div><p class="ohl">Incassato ad Oggi</p><p class="ohv" style="color:#2E8B5F">{fe(ON["incassato"])}</p><p class="ohs">{pctO}% del target</p></div>'
    f'<div><p class="ohl">Gap Residuo Oggi</p><p class="ohv" style="color:#C0392B">{fe(ON["deltaOggi"])}</p><p class="ohs">Da recuperare</p></div>'
    f'<div><p class="ohl">Previsto a Dicembre</p><p class="ohv" style="color:#D97706">{fe(ON["previstoDic"])}</p><p class="ohs">Incassi certi 2026</p></div>'
    f'<div><p class="ohl">Gap a Fine Anno</p><p class="ohv" style="color:#C0392B">{fe(ON["deltaDic"])}</p><p class="ohs">Con dati attuali</p></div>'
    f'<div>'
    f'<p class="ohl" style="margin-bottom:6px">Progressione</p>'
    f'<div style="background:#FAF6EE;border-radius:4px;height:9px;overflow:hidden;margin-bottom:3px">'
    f'<div style="height:100%;width:{min(pctO,100)}%;background:linear-gradient(90deg,#C8A951,#E8CC7A);border-radius:4px"></div></div>'
    f'<div style="display:flex;justify-content:space-between;font-size:.59rem;color:#64748B">'
    f'<span>Inc: <strong style="color:#0B1E3D">{pctO}%</strong></span>'
    f'<span>Prev: <strong style="color:#D97706">{pctP}%</strong></span>'
    f'<span>Target: 100%</span></div>'
    f'<div style="background:#FAF6EE;border-radius:4px;height:5px;overflow:hidden;margin-top:4px;opacity:.55">'
    f'<div style="height:100%;width:{min(pctP,100)}%;background:linear-gradient(90deg,#D97706,#F59E0B);border-radius:4px"></div></div>'
    f'<p style="font-size:.6rem;color:#64748B;margin-top:4px">Gap: {fe(ON["deltaDic"])}</p>'
    f'</div>'
)

# Ranking
medals = ["🥇", "🥈", "🥉"] + [f"{i}°" for i in range(4, 11)]
ranked = sorted(
    [(c['name'], KB.get(c['name'], {})) for c in D['collaboratori']],
    key=lambda x: -x[1].get('premioAnnuo', 0)
)[:10]
rank_rows = ""
for i, (nm, kb) in enumerate(ranked):
    if not kb.get('polizze', 0) and not kb.get('premioAnnuo', 0):
        continue
    cvp = round(kb.get('conv', 0) * 100)
    tc = "tg" if cvp >= 80 else "ta" if cvp >= 50 else "tr2"
    rank_rows += (
        f'<tr><td>{medals[i]}</td><td>{nm}</td>'
        f'<td>{kb.get("apptTot",0)}</td>'
        f'<td><strong>{kb.get("polizze",0)}</strong></td>'
        f'<td><strong class="num">{fe(kb.get("premioAnnuo",0))}</strong></td>'
        f'<td class="num">{fe(kb.get("premiIncassati",0))}</td>'
        f'<td>{tag(tc, str(cvp)+"%")}</td></tr>'
    )

# Prodotti
prod_map = {}
for r in D['giornalieri']:
    if r['esito'] == 'Sottoscritto' and r['tipoPol']:
        tp = r['tipoPol'].replace('Protezione Casa e Famiglia', 'Casa+Famiglia').replace('Protezione ', '')
        prod_map.setdefault(tp, {'n': 0, 'pa': 0.0})
        prod_map[tp]['n'] += 1
        prod_map[tp]['pa'] += r['premioAnnuo']

prods = sorted(prod_map.items(), key=lambda x: -x[1]['n'])
maxN = prods[0][1]['n'] if prods else 1
totP = sum(v['n'] for _, v in prods)
PCOLS = ["linear-gradient(90deg,#0B1E3D,#1E3A6E)", "linear-gradient(90deg,#1E3A6E,#2A5299)",
         "linear-gradient(90deg,#C8A951,#E8CC7A)", "linear-gradient(90deg,#2E8B5F,#3BA870)",
         "linear-gradient(90deg,#D97706,#F59E0B)", "#7C3AED", "#94A3B8"]
prod_html = "<p class='bcht'>&#x1F6CD;&#xFE0F; Mix Prodotti &mdash; Distribuzione Polizze</p>"
for i, (nm, v) in enumerate(prods):
    pct = round(v['n'] / max(totP, 1) * 100)
    w = round(v['n'] / maxN * 100)
    prod_html += (
        f'<div class="brr">'
        f'<span class="brl">{nm}</span>'
        f'<div class="bro"><div class="bri" style="width:{w}%;background:{PCOLS[i % len(PCOLS)]}">'
        f'<span>{v["n"]} pol.</span></div></div>'
        f'<span class="brn">{pct}% &middot; {fe(v["pa"])}</span></div>'
    )

# Top 3 / Top 5 rank cards
def rank_card(title, key, eur=True, n=5):
    items = sorted(KB.items(), key=lambda x: -x[1].get(key, 0))[:n]
    rows = ""
    for i, (nm, kb) in enumerate(items):
        v = fe(kb.get(key, 0)) if eur else fn(kb.get(key, 0))
        rows += (
            f'<div class="rki"><span class="rip">{medals[i]}</span>'
            f'<span class="rin">{nm}</span>'
            f'<div><div class="riv num">{v}</div>'
            f'<div class="ris">{kb.get("polizze",0)} pol.</div></div></div>'
        )
    return f'<div class="rkc"><div class="rkh">{title}</div>{rows}</div>'


top3 = (rank_card("&#x1F4C5; Top5 &middot; Appt", "apptTot", False)
        + rank_card("&#x1F4B6; Top5 &middot; Premio", "premioAnnuo", True)
        + rank_card("&#x1F4BC; Top5 &middot; Incassati", "premiIncassati", True))

# FB Summary cards
nG = G.get('fbAttivi', sum(1 for k in KB.values() if k['statoAttivo'] == 'green'))
nA = sum(1 for k in KB.values() if k['statoAttivo'] == 'amber')
nR = sum(1 for k in KB.values() if k['statoAttivo'] == 'red')
totPA = sum(k.get('premioAnnuo', 0) for k in KB.values())
fbsum = (
    card("FB Attivi &#x25CF; Verde (pol&ge;4)", str(nG), "Raggiungono il minimo mensile", "green", "&#x2705;")
    + card("FB Quasi Attivi &#x25CF; Arancione", str(nA), "Polizze tra 2 e 3", "amb", "&#x26A0;&#xFE0F;")
    + card("FB Non Attivi &#x25CF; Rosso (pol&lt;2)", str(nR), "Colloquio urgente", "red", "&#x274C;")
    + card("Premio Annuo Totale Team", fe(totPA), f"{G['polizze']} polizze YTD", "", "&#x1F4B0;")
)

# FB Table rows
fb_rows = ""
for c in D['collaboratori']:
    nm = c['name']
    kb = KB.get(nm, {})
    rv = RL.get(nm, {})
    actv = kb.get('statoAttivo', 'neutral')
    dlt = kb.get('deltaObj', 0)
    dtag = tag("tg", f"+{dlt}") if dlt >= 0 else tag("ta", str(dlt)) if dlt == -1 else tag("tr2", str(dlt))
    cb = kb.get('callback', 0)
    cbtag = tag("ta", str(cb)) if cb > 0 else f'<span style="color:#94A3B8">0</span>'
    proj = fe(rv['proiezione']) if rv.get('proiezione', 0) > 0 else "&#x2014;"
    proj_col = "#2E8B5F" if "↑" in rv.get('stato', '') else "#64748B"
    # Lavorazione data for this FB
    fb_lav = [p for p in PL_data if p['fb'] == nm]
    n_lav_fb = len(fb_lav)
    pf_lav_fb = sum(p['premioFirma'] for p in fb_lav)
    lav_cell = f'<span class="num" style="color:#D97706;font-weight:600">{n_lav_fb}</span>' if n_lav_fb > 0 else '<span style="color:#94A3B8">0</span>'
    pf_lav_cell = f'<span class="num" style="color:#D97706">{fe(pf_lav_fb)}</span>' if pf_lav_fb > 0 else '<span style="color:#94A3B8">&#x2014;</span>'
    # Bigger dot
    dot_big = (f'<span style="font-size:1.8rem;line-height:1;color:#2E8B5F">&#x25CF;</span>' if actv == 'green' else
               f'<span style="font-size:1.8rem;line-height:1;color:#D97706">&#x25CF;</span>' if actv == 'amber' else
               f'<span style="font-size:1.8rem;line-height:1;color:#C0392B">&#x25CF;</span>' if actv == 'red' else
               f'<span style="font-size:1.8rem;line-height:1;color:#94A3B8">&#x25CF;</span>')
    fb_rows += (
        f'<tr>'
        f'<td>{nm}</td>'
        f'<td style="font-size:.67rem;color:#64748B">{c["fbo"]}</td>'
        f'<td class="num" style="text-align:center">{kb.get("apptTot",0)}</td>'
        f'<td style="text-align:center"><strong class="num">{kb.get("polizze",0)}</strong></td>'
        f'<td style="text-align:center">{dot_big}</td>'
        f'<td class="num" style="text-align:center">{fe(kb.get("premioAnnuo",0))}</td>'
        f'<td class="num" style="text-align:center">{fe(kb.get("premiIncassati",0))}</td>'
        f'<td style="text-align:center">{lav_cell}</td>'
        f'<td style="text-align:center">{pf_lav_cell}</td>'
        f'<td style="text-align:center">{cbtag}</td>'
        f'</tr>'
    )

top5 = (rank_card("&#x1F4C5; Top5 Appt", "apptTot", False)
        + rank_card("&#x1F4B6; Top5 Premio", "premioAnnuo", True)
        + rank_card("&#x1F4BC; Top5 Incassati", "premiIncassati", True))


# ── OBJ SETTIMANALE ────────────────────────────────────────────────
# Calcola % incassata settimanale approssimata
inizio_anno = _dt(2026, 1, 1)
oggi_dt = _dt.today()
sett_tot = 52
sett_pass = min(int((oggi_dt - inizio_anno).days / 7) + 1, 52)
# Target progressivo settimanale
target_sett_fps   = round(G['target'] * sett_pass / sett_tot, 2)
target_sett_on    = round(ON['obiettivo'] * sett_pass / sett_tot, 2)
pct_fps_sett      = round(G['premiIncassati'] / max(G['target'], 1) * 100, 1)
pct_on_sett       = round(ON['incassato'] / max(ON['obiettivo'], 1) * 100, 1)
pct_prog_fps      = round(G['premiIncassati'] / max(target_sett_fps, 1) * 100, 1)
pct_prog_on       = round(ON['incassato'] / max(target_sett_on, 1) * 100, 1)
gap_sett_fps      = round(target_sett_fps - G['premiIncassati'], 2)
gap_sett_on       = round(target_sett_on - ON['incassato'], 2)

obj_html = f'''<div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px">
  <div class="card gold">
    <p class="cl">&#x1F3AF; Obiettivo Personale FPS &mdash; Settimana {sett_pass}</p>
    <p class="cv num">{pct_fps_sett}%</p>
    <p class="csub">Incassato {fe(G["premiIncassati"])} su target {fe(G["target"])}</p>
    <div style="margin-top:10px">
      <div style="display:flex;justify-content:space-between;font-size:.7rem;color:var(--mut);margin-bottom:4px">
        <span>Target progressivo sett. {sett_pass}</span><span class="num">{fe(target_sett_fps)}</span>
      </div>
      <div style="background:var(--cream);border-radius:4px;height:10px;overflow:hidden">
        <div style="height:100%;width:{min(pct_prog_fps,100)}%;background:{"linear-gradient(90deg,var(--gr),var(--gr2))" if pct_prog_fps>=100 else "linear-gradient(90deg,var(--amb),#F59E0B)" if pct_prog_fps>=70 else "linear-gradient(90deg,var(--red),#E74C3C)"};border-radius:4px"></div>
      </div>
      <div style="display:flex;justify-content:space-between;font-size:.7rem;margin-top:4px">
        <span style="color:var(--mut)">{pct_prog_fps}% del target progressivo</span>
        <span class="num" style="color:{"var(--gr)" if gap_sett_fps<=0 else "var(--red)"}">{"✅ In linea" if gap_sett_fps<=0 else f"Gap: {fe(gap_sett_fps)}"}</span>
      </div>
    </div>
    <div style="margin-top:12px;display:grid;grid-template-columns:repeat(4,1fr);gap:8px;text-align:center">
      {"".join(f'<div style="background:var(--cream);border-radius:6px;padding:8px 4px"><p style="font-size:.58rem;color:var(--mut);text-transform:uppercase;letter-spacing:.05em">Sett.{i+1}</p><p class="num" style="font-size:.85rem;color:var(--navy)">{fe(G["target"]*( (i+1)/52))}</p><p style="font-size:.58rem;color:var(--mut)">target</p></div>' for i in range(min(sett_pass,4)))}
    </div>
  </div>
  <div class="card" style="border-top-color:var(--gold)">
    <p class="cl">&#x1F3AF; Obiettivo Gruppo Onorato &mdash; Settimana {sett_pass}</p>
    <p class="cv num">{pct_on_sett}%</p>
    <p class="csub">Incassato {fe(ON["incassato"])} su target {fe(ON["obiettivo"])}</p>
    <div style="margin-top:10px">
      <div style="display:flex;justify-content:space-between;font-size:.7rem;color:var(--mut);margin-bottom:4px">
        <span>Target progressivo sett. {sett_pass}</span><span class="num">{fe(target_sett_on)}</span>
      </div>
      <div style="background:var(--cream);border-radius:4px;height:10px;overflow:hidden">
        <div style="height:100%;width:{min(pct_prog_on,100)}%;background:{"linear-gradient(90deg,var(--gr),var(--gr2))" if pct_prog_on>=100 else "linear-gradient(90deg,var(--amb),#F59E0B)" if pct_prog_on>=70 else "linear-gradient(90deg,var(--red),#E74C3C)"};border-radius:4px"></div>
      </div>
      <div style="display:flex;justify-content:space-between;font-size:.7rem;margin-top:4px">
        <span style="color:var(--mut)">{pct_prog_on}% del target progressivo</span>
        <span class="num" style="color:{"var(--gr)" if gap_sett_on<=0 else "var(--red)"}">{"✅ In linea" if gap_sett_on<=0 else f"Gap: {fe(gap_sett_on)}"}</span>
      </div>
    </div>
  </div>
</div>'''

# ── CHALLENGE HTML ──────────────────────────────────────────────────
medals_ch = ['&#x1F947;','&#x1F948;','&#x1F949;'] + [f'{i}&#xB0;' for i in range(4,20)]
ch_classifica = ""
for i,r in enumerate(ch_rows):
    pa_v = round(float(r['pa']),2) if r['pa'] else 0
    pol_v = r['pol'] or '0'
    is_winner = (int(r['pol'] or 0) >= int(ch_min_pol or 0) if ch_min_pol else False)
    ch_classifica += f'<tr style="{"background:rgba(200,169,81,.08)" if is_winner else ""}"><td>{medals_ch[i]}</td><td><strong>{r["name"]}</strong>{"&nbsp;&#x1F3C6;" if is_winner else ""}</td><td class="num">{pol_v}</td><td class="num">{fe(pa_v)}</td></tr>'

ch_part_html = "".join(f'<span style="display:inline-block;background:var(--cream);border-radius:6px;padding:4px 10px;font-size:.78rem;margin:3px">{p}</span>' for p in ch_part)

# Obiettivi medaglia
ch_obj_html = ""
if ch_min_pol:
    ch_obj_html += f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px"><span style="font-size:1.2rem">&#x1F4CB;</span><span style="font-size:.82rem"><strong>Polizze:</strong> &ge; {ch_min_pol}</span></div>'
if ch_min_pa:
    ch_obj_html += f'<div style="display:flex;align-items:center;gap:8px"><span style="font-size:1.2rem">&#x1F4B6;</span><span style="font-size:.82rem"><strong>Premio Annuo:</strong> &ge; {fe(float(ch_min_pa) if ch_min_pa else 0)}</span></div>'

challenge_html = f'''<div class="onc" style="border-top-color:#7C3AED">
  <div class="onh"><h3>&#x1F3C1; Challenge Attiva &mdash; {ch_inizio_str} / {ch_fine_str}</h3><span>Fonte: Foglio Challenge</span></div>
  <div style="padding:18px 20px">
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:20px">
      <div>
        <div class="tw" style="margin-bottom:0">
          <div class="twh"><h3>&#x1F3C6; Classifica</h3><span>Solo partecipanti inclusi</span></div>
          <table><thead><tr><th>#</th><th>Family Banker</th><th>Polizze</th><th>Premio Annuo</th></tr></thead>
          <tbody>{ch_classifica}</tbody></table>
        </div>
      </div>
      <div>
        <div class="card" style="margin-bottom:14px;border-top-color:#7C3AED">
          <p class="cl">&#x1F3AF; Obiettivi per vincere la medaglia</p>
          <div style="margin-top:10px">{ch_obj_html}</div>
        </div>
        <div class="card" style="border-top-color:#7C3AED">
          <p class="cl">&#x1F465; Partecipanti ({len(ch_part)})</p>
          <div style="margin-top:8px">{ch_part_html}</div>
        </div>
      </div>
    </div>
  </div>
</div>'''



# Trend tables
hdr_mesi = "".join(f"<th>{m}</th>" for m in MB)
tPol = pol_mese
tApp = [sum((AM.get(fb, {}).get(m, 0)) for fb in FBS) for m in mesi_raw]

trend_pol_rows = ""
for fb in FBS:
    obj = next((c['objMese'] for c in D['collaboratori'] if c['name'] == fb), 0)
    tot = sum(PM.get(fb, {}).get(m, 0) for m in mesi_raw)
    cells = ""
    for m in mesi_raw:
        v = PM.get(fb, {}).get(m, 0)
        cl2 = "ck0" if v == 0 else "ckok" if (obj > 0 and v >= obj) else "ckw" if (obj > 0 and v >= obj - 1) else "ckr"
        cells += f'<td><span class="ck {cl2}">{v or "&#x2014;"}</span></td>'
    trend_pol_rows += f'<tr><td>{fb}</td>{cells}<td><strong>{tot or "&#x2014;"}</strong></td></tr>'
trend_pol_rows += (
    '<tr class="totr"><td>Totale Team</td>'
    + "".join(f'<td><strong>{v}</strong></td>' for v in tPol)
    + f'<td><strong>{sum(tPol)}</strong></td></tr>'
)

trend_appt_rows = ""
for fb in FBS:
    tot = sum(AM.get(fb, {}).get(m, 0) for m in mesi_raw)
    cells = "".join(
        f'<td><span class="ck {"ckok" if AM.get(fb,{}).get(m,0)>0 else "ck0"}">{AM.get(fb,{}).get(m,0) or "&#x2014;"}</span></td>'
        for m in mesi_raw)
    trend_appt_rows += f'<tr><td>{fb}</td>{cells}<td><strong>{tot or "&#x2014;"}</strong></td></tr>'
trend_appt_rows += (
    '<tr class="totr"><td>Totale Team</td>'
    + "".join(f'<td><strong>{v}</strong></td>' for v in tApp)
    + f'<td><strong>{sum(tApp)}</strong></td></tr>'
)

def _tcard(i):
    return (
        f"<div class='card gold'>"
        f"<div class='cico'>&#x1F4C5;</div>"
        f"<p class='cl'>{MB[i]}</p>"
        f"<p style='font-family:Outfit,sans-serif;font-size:1.25rem;font-weight:700;color:var(--navy);line-height:1.2;margin-bottom:4px'>{fe(pf_mese[i])} <span style='font-size:.65rem;color:var(--mut);font-weight:400'>Firma</span></p>"
        f"<p style='font-family:Outfit,sans-serif;font-size:1.25rem;font-weight:700;color:var(--gr);line-height:1.2;margin-bottom:6px'>{fe(pi_mese[i])} <span style='font-size:.65rem;color:var(--mut);font-weight:400'>Inc.</span></p>"
        f"<p style='font-family:Outfit,sans-serif;font-size:.92rem;font-weight:600;color:var(--mut)'>{tPol[i]} pol &middot; {tApp[i]} appt</p>"
        f"</div>"
    )
tsum = "".join(_tcard(i) for i in range(len(mesi_raw)))

# Azioni
critici = [nm for nm, kb in KB.items() if kb.get('statoAttivo') == 'red']
zeroPol = [nm for nm, kb in KB.items() if kb.get('polizze', 0) == 0 and kb.get('statoAttivo') != 'neutral']
top_perf = sorted(
    [(nm, kb) for nm, kb in KB.items() if kb.get('statoAttivo') == 'green' and kb.get('polizze', 0) >= 4],
    key=lambda x: -x[1].get('premioAnnuo', 0))[:5]

alerts = (
    f'<div class="alert d"><p class="alt">&#x1F534; Ritmo Insufficiente &mdash; Gap {fe(G["target"]-G["premiIncassati"])}</p>'
    f'<div class="alb">Incassato {fe(G["premiIncassati"])} su {fe(G["target"])} ({round(G["premiIncassati"]/G["target"]*100)}%). '
    f'<ul><li>~13 polizze/settimana per recuperare</li><li>Richiamare i {G["callback"]} callback aperti</li></ul></div></div>'

    + f'<div class="alert d"><p class="alt">&#x1F534; {len(critici)} Family Banker Non Attivi (&lt;2 polizze)</p>'
    f'<div class="alb"><ul>' + "".join(f"<li><strong>{n}</strong></li>" for n in critici[:8]) + "</ul></div></div>"

    + f'<div class="alert w"><p class="alt">&#x1F7E1; {G["callback"]} Callback Aperti</p>'
    f'<div class="alb">Richiamare entro venerd&igrave;. Conversione media: {fp(G["convRate"])}.</div></div>'

    + f'<div class="alert w"><p class="alt">&#x1F7E1; {len(zeroPol)} FB Senza Polizze nel 2026</p>'
    f'<div class="alb"><ul>' + "".join(f"<li><strong>{n}</strong></li>" for n in zeroPol) + "</ul></div></div>"

    + f'<div class="alert ok"><p class="alt">&#x1F7E2; Top Performer</p>'
    f'<div class="alb"><ul>' + "".join(f"<li><strong>{n}</strong>: {kb['polizze']} pol., {fe(kb['premioAnnuo'])}</li>" for n, kb in top_perf) + "</ul></div></div>"

    + f'<div class="alert i"><p class="alt">&#x1F535; {len(PL_data)} Polizze in Lavorazione</p>'
    f'<div class="alb"><ul>' + "".join(f"<li><strong>{p['fb']}</strong> &mdash; {p['cliente']} &mdash; {fe(p['premioAnnuo'])}</li>" for p in PL_data) + "</ul></div></div>"
)

delta_rows = ""
for c in D['collaboratori']:
    kb = KB.get(c['name'], {})
    pol = kb.get('polizze', 0)
    actv = kb.get('statoAttivo', 'neutral')
    stag = (tag("tg", "&#x25CF; Attivo") if actv == "green" else
            tag("ta", "&#x25CF; Quasi") if actv == "amber" else
            tag("tr2", "&#x25CF; Non attivo") if actv == "red" else
            tag("tn", "&#x2014;"))
    gap = pol - MESE_CORR
    gcol = "#2E8B5F" if gap >= 0 else "#D97706" if gap >= -2 else "#C0392B"
    delta_rows += f'<tr><td>{c["name"]}</td><td class="num">{pol}</td><td>{c["objMese"]}/m</td><td>{stag}</td><td class="num" style="color:{gcol}">{gap}</td></tr>'

piano_rows = (
    f'<tr><td>{tag("tr2","&#x1F534; Critica")}</td><td>Richiamare i {G["callback"]} callback aperti</td><td>Tutti i FB</td><td>Entro venerd&igrave;</td></tr>'
    f'<tr><td>{tag("tr2","&#x1F534; Critica")}</td><td>Colloquio con i {len(critici)} FB non attivi</td><td>FPS</td><td>Questa settimana</td></tr>'
    f'<tr><td>{tag("ta","&#x1F7E1; Alta")}</td><td>Processare {len(PL_data)} polizze in lavorazione</td><td>Backoffice</td><td>Entro 48h</td></tr>'
    f'<tr><td>{tag("ta","&#x1F7E1; Alta")}</td><td>Condivisione best practice top performer</td><td>FPS</td><td>Prossima riunione</td></tr>'
    f'<tr><td>{tag("tb","&#x1F535; Media")}</td><td>Training Protezione Salute</td><td>Tutto il team</td><td>Fine mese</td></tr>'
)

# FB select options
fb_options = "\n".join(
    f'<option value="{c["name"]}">{c["name"]}</option>'
    for c in sorted(D['collaboratori'], key=lambda c: c['name'])
)

# Colloquio schede
def colloquio_html(c):
    nm = c['name']
    kb = KB.get(nm, {})
    rv = RL.get(nm, {})
    pols = PZ.get(nm, [])
    plav = [p for p in PL_data if p['fb'] == nm]
    cb = kb.get('callback', 0)
    actv = kb.get('statoAttivo', 'neutral')
    atag = (tag("tg", "&#x25CF; Attivo") if actv == "green" else
            tag("ta", "&#x25CF; Quasi attivo") if actv == "amber" else
            tag("tr2", "&#x25CF; Non attivo") if actv == "red" else
            tag("tn", "Obj = 0"))
    conv = round(kb.get('polizze', 0) / max(kb.get('apptTot', 0), 1) * 100) if kb.get('apptTot', 0) else 0
    pO = kb.get('premiIncassati', 0) / max(c['objPremio'], 1) if c['objPremio'] else 0
    ini = "".join(w[0] for w in nm.split())[:2].upper()
    dlt = kb.get('deltaObj', 0)
    pol = kb.get('polizze', 0)
    pc_col = "#2E8B5F" if pO >= 0.5 else "#D97706" if pO >= 0.2 else "#C0392B"
    rv_col = "#2E8B5F" if "↑" in rv.get('stato', '') else "#D97706"
    pm_val = kb.get('premioAnnuo', 0) / max(pol, 1) if pol else 0

    pol_cells = "".join(
        f'<td><span class="ck {"ck0" if PM.get(nm,{}).get(m,0)==0 else "ckok" if (c["objMese"]>0 and PM.get(nm,{}).get(m,0)>=c["objMese"]) else "ckw"}">'
        f'{PM.get(nm,{}).get(m,0) or "&#x2014;"}</span></td>'
        for m in mesi_raw)
    ap_cells = "".join(
        f'<td><span class="ck {"ckok" if AM.get(nm,{}).get(m,0)>0 else "ck0"}">'
        f'{AM.get(nm,{}).get(m,0) or "&#x2014;"}</span></td>'
        for m in mesi_raw)

    pols_html = ""
    if pols:
        rows = "".join(
            f'<tr><td>{p.get("dataPolizza","&#x2014;")}</td>'
            f'<td><strong>{p.get("cliente","&#x2014;")}</strong></td>'
            f'<td>{p.get("tipoPol","&#x2014;")}</td>'
            f'<td class="num">{fe(p.get("premioFirma",0))}</td>'
            f'<td class="num">{fe(p.get("premioAnnuo",0))}</td>'
            f'<td>{p.get("frazionamento","&#x2014;")}</td>'
            f'<td>{tag("tg" if p.get("statoPolizza")=="Processata" else "tb" if p.get("statoPolizza")=="In lavorazione" else "tr2", p.get("statoPolizza","&#x2014;"))}</td></tr>'
            for p in pols)
        pols_html = (
            f'<div class="csec"><p class="csect">&#x1F4CB; Polizze Sottoscritte ({len(pols)})</p>'
            f'<table class="mt"><thead><tr><th>Data</th><th>Cliente</th><th>Tipo</th><th>P.Firma</th><th>P.Annuo</th><th>Fraz.</th><th>Stato</th></tr></thead>'
            f'<tbody>{rows}</tbody></table></div>'
        )

    lav_html = ""
    if plav:
        rows = "".join(
            f'<tr><td><strong>{p["cliente"]}</strong></td><td>{p["tipoPol"]}</td>'
            f'<td class="num">{fe(p["premioFirma"])}</td><td class="num">{fe(p["premioAnnuo"])}</td></tr>'
            for p in plav)
        lav_html = (
            f'<div class="csec"><p class="csect">&#x2699;&#xFE0F; In Lavorazione ({len(plav)})</p>'
            f'<table class="mt"><thead><tr><th>Cliente</th><th>Tipo</th><th>P.Firma</th><th>P.Annuo</th></tr></thead>'
            f'<tbody>{rows}</tbody></table></div>'
        )

    pts = []
    if pol == 0 and c['objMese'] > 0:
        pts.append(("&#x1F6A8;", "<strong>Nessuna polizza nel 2026.</strong> Colloquio urgente."))
    elif actv == 'green':
        pts.append(("&#x2705;", f"<strong>FB Attivo</strong>: {pol} polizze &mdash; sopra il minimo mensile."))
    elif actv == 'amber':
        pts.append(("&#x26A0;&#xFE0F;", f"<strong>Quasi attivo</strong>: {pol} polizze. Basta 1-2 per rientrare."))
    else:
        pts.append(("&#x1F4C9;", f"<strong>Non attivo</strong>: {pol} polizze. Target: &ge;4 entro fine mese."))
    if kb.get('apptTot', 0) > 2 and conv < 30:
        pts.append(("&#x26A0;&#xFE0F;", f"<strong>Conversione bassa ({conv}%)</strong>: rivedere tecnica di chiusura."))
    if cb > 0:
        pts.append(("&#x1F4DE;", f"<strong>{cb} callback aperti</strong>: richiamare con priorit&agrave;."))
    if pol > 0 and pm_val < 500:
        pts.append(("&#x1F4A1;", f"<strong>Premio medio basso</strong> ({fe(pm_val)}/pol.): up-sell Protezione Salute."))
    if c['objPremio'] > 0 and pO > 0.5:
        pts.append(("&#x1F4C8;", f"<strong>Buon avanzamento sul budget</strong>: {round(pO*100)}% raggiunto."))
    azione = (
        "Colloquio urgente: piano settimanale con obiettivi e affiancamento." if pol == 0
        else f"Aumentare da {kb.get('apptTot',0)} a {max(kb.get('apptTot',0)+2,4)} appt/mese." if conv >= 100
        else f"Recuperare {abs(dlt)} polizze: callback e prospect caldi." if dlt < -4
        else "Mantenere il ritmo e cercare referral dai clienti soddisfatti."
    )
    pts.append(("&#x1F4A1;", f"<strong>Azione consigliata:</strong> {azione}"))
    anl = "".join(f'<div class="anlr"><span>{i}</span><span class="anlt">{t}</span></div>' for i, t in pts)

    return f"""<div class="collg">
      <div class="cpf">
        <div class="cpav">{ini}</div>
        <p class="cpnm">{nm}</p>
        <p class="cpfb">{c['fbo']} &middot; {c['gruppo']}</p>
        <div style="margin-bottom:9px">{atag}</div>
        <div class="cpst" style="flex-direction:column;align-items:flex-start;gap:2px"><span class="csl">Email</span><span class="csv" style="font-size:.58rem;word-break:break-all;text-align:left">{c['email'] or '&#x2014;'}</span></div>
        <div class="cpst"><span class="csl">Cellulare</span><span class="csv">{c.get('cellulare','') or '&#x2014;'}</span></div>
        <div class="cpst"><span class="csl">Obj pol./mese</span><span class="csv">{c['objMese']}</span></div>
        <div class="cpst"><span class="csl">Delta vs target</span><span class="csv" style="color:{'#2E8B5F' if dlt>=0 else '#D97706' if dlt==-1 else '#C0392B'}">{dlt}</span></div>
        <div class="cpst"><span class="csl">Obj premio</span><span class="csv">{fe(c['objPremio']) if c['objPremio'] else '&#x2014;'}</span></div>
        <div class="cpst"><span class="csl">Callback</span><span class="csv" style="color:{'#D97706' if cb>0 else '#2E8B5F'}">{cb}</span></div>
        <div class="cpst"><span class="csl">Provvigioni</span><span class="csv">{fe(kb.get('provv',0))}</span></div>
      </div>
      <div>
        <div class="ckg">
          <div class="ckc gold"><p class="ckl">Appt Com.</p><p class="ckv num">{kb.get('apptTot',0)}</p><p class="cks">YTD</p></div>
          <div class="ckc"><p class="ckl">Polizze YTD</p><p class="ckv num">{pol}</p><p class="cks">Conv. {conv}%</p></div>
          <div class="ckc green"><p class="ckl">Premio Annuo</p><p class="ckv num" style="font-size:.95rem">{fe(kb.get('premioAnnuo',0))}</p><p class="cks">Med. {fe(round(pm_val))}</p></div>
          <div class="ckc amb"><p class="ckl">Premi Incassati</p><p class="ckv num" style="font-size:.95rem">{fe(kb.get('premiIncassati',0))}</p><p class="cks">{round(kb.get('premiIncassati',0)/max(kb.get('premioAnnuo',1),1)*100)}% p.a.</p></div>
          <div class="ckc {'green' if pO>=0.5 else 'amb' if pO>=0.2 else 'red'}"><p class="ckl">% Obj Premio</p><p class="ckv num" style="color:{pc_col}">{round(pO*100)}%</p><p class="cks">{fe(c['objPremio']) if c['objPremio'] else 'no obj'}</p></div>
          <div class="ckc"><p class="ckl">Proiezione Dic.</p><p class="ckv num" style="font-size:.95rem;color:{rv_col}">{fe(rv['proiezione']) if rv.get('proiezione',0)>0 else '&#x2014;'}</p><p class="cks">{rv.get('stato','&#x2014;')}</p></div>
        </div>
        <div class="csec"><p class="csect">&#x1F4C8; Andamento Mensile</p>
          <table class="mt"><thead><tr><th></th>{''.join(f"<th>{m}</th>" for m in MB)}<th>Tot.</th></tr></thead>
          <tbody>
            <tr><td>Pol.</td>{pol_cells}<td><strong>{len(pols)}</strong></td></tr>
            <tr><td>Appt</td>{ap_cells}<td><strong>{sum(AM.get(nm,{}).get(m,0) for m in mesi_raw)}</strong></td></tr>
          </tbody></table>
        </div>
        {pols_html}{lav_html}
        <div class="anl"><h4>&#x1F4CB; Analisi per il Colloquio</h4>{anl}</div>
      </div>
    </div>"""


coll_data = {c['name']: colloquio_html(c) for c in D['collaboratori']}
coll_json = json.dumps(coll_data, ensure_ascii=True)

# ─── DATA AGGIORNAMENTO ──────────────────────────────────────────────
oggi = datetime.today().strftime('%d/%m/%Y')

# ─── CSS ────────────────────────────────────────────────────────────

# ─── CALCOLI OBJ SETTIMANALE E CHALLENGE ─────────────────────────────
import re as _re2, datetime as _dt2
from collections import defaultdict as _dd2
import math as _math2

# Carica xl se non già disponibile (alias)
_xl = xl  # già caricato sopra

# Weekly data dal Dati Giornalieri
_weekly = _dd2(lambda: {"pa": 0.0, "pol": 0})
for _r in D["giornalieri"]:
    if _r["esito"] != "Sottoscritto":
        continue
    # La data è già una stringa dd/mm/yyyy o un oggetto
    _dp = _r.get("dataPolizza", "")
    if not _dp:
        continue
    try:
        if isinstance(_dp, str) and len(_dp) >= 10:
            _d = _dt2.datetime.strptime(_dp[:10], "%d/%m/%Y") if "/" in _dp else _dt2.datetime.strptime(_dp[:10], "%Y-%m-%d")
        else:
            continue
        _wk = _d.isocalendar()[1]
        _weekly[f"S.{_wk}"]["pa"] += _r["premioAnnuo"]
        _weekly[f"S.{_wk}"]["pol"] += 1
    except:
        pass

_sett_ord = sorted(_weekly.keys(), key=lambda x: int(_re2.search(r"\d+", x).group()))
_tot_pa_w = sum(w["pa"] for w in _weekly.values())
_tot_inc  = G["premiIncassati"]
_tot_inc_on = ON["incassato"]
_target_300 = G["target"]
_target_180 = ON["obiettivo"]

obj_rows_300 = ""
obj_rows_180 = ""
_cum_300 = 0.0
_cum_180 = 0.0
for _s in _sett_ord:
    _w = _weekly[_s]
    _frac = _w["pa"] / max(_tot_pa_w, 1)
    _i300 = round(_frac * _tot_inc, 2)
    _i180 = round(_frac * _tot_inc_on, 2)
    _cum_300 += _i300
    _cum_180 += _i180
    _p300 = round(_cum_300 / max(_target_300, 1) * 100, 1)
    _p180 = round(_cum_180 / max(_target_180, 1) * 100, 1)
    _c300 = "#2E8B5F" if _p300 >= 25 else "#D97706" if _p300 >= 10 else "#C0392B"
    _c180 = "#2E8B5F" if _p180 >= 25 else "#D97706" if _p180 >= 10 else "#C0392B"
    _ps300 = round(_i300 / max(_target_300, 1) * 100, 2)  # % settimana singola
    _ps180 = round(_i180 / max(_target_180, 1) * 100, 2)
    _cs300 = "#2E8B5F" if _ps300 >= 1 else "#D97706" if _ps300 >= 0.5 else "#C0392B"
    _cs180 = "#2E8B5F" if _ps180 >= 1 else "#D97706" if _ps180 >= 0.5 else "#C0392B"
    obj_rows_300 += (
        f"<tr><td>{_s}</td><td class='num'>{_w['pol']}</td>"
        f"<td class='num'>{fe(_i300)}</td>"
        f"<td><span class='num' style='font-size:.72rem;color:{_cs300};font-weight:700'>{_ps300}%</span></td>"
        f"<td class='num'>{fe(_cum_300)}</td>"
        f"<td><div style='display:flex;align-items:center;gap:5px'>"
        f"<div style='flex:1;background:#FAF6EE;border-radius:3px;height:7px;overflow:hidden'>"
        f"<div style='height:100%;width:{min(_p300,100)}%;background:{_c300};border-radius:3px'></div></div>"
        f"<span class='num' style='font-size:.7rem;color:{_c300};font-weight:700'>{_p300}%</span>"
        f"</div></td></tr>"
    )
    obj_rows_180 += (
        f"<tr><td>{_s}</td><td class='num'>{_w['pol']}</td>"
        f"<td class='num'>{fe(_i180)}</td>"
        f"<td><span class='num' style='font-size:.72rem;color:{_cs180};font-weight:700'>{_ps180}%</span></td>"
        f"<td class='num'>{fe(_cum_180)}</td>"
        f"<td><div style='display:flex;align-items:center;gap:5px'>"
        f"<div style='flex:1;background:#FAF6EE;border-radius:3px;height:7px;overflow:hidden'>"
        f"<div style='height:100%;width:{min(_p180,100)}%;background:{_c180};border-radius:3px'></div></div>"
        f"<span class='num' style='font-size:.7rem;color:{_c180};font-weight:700'>{_p180}%</span>"
        f"</div></td></tr>"
    )

tot_pol_obj = sum(_w["pol"] for _w in _weekly.values())
fe_tot_inc_300 = fe(_tot_inc)
fe_tot_inc_180 = fe(_tot_inc_on)
fe_target_300  = fe(_target_300)
fe_target_180  = fe(_target_180)
pct_fin_300 = round(_tot_inc / max(_target_300, 1) * 100, 1)
pct_fin_180 = round(_tot_inc_on / max(_target_180, 1) * 100, 1)
col_fin_300 = "#2E8B5F" if pct_fin_300 >= 25 else "#D97706" if pct_fin_300 >= 10 else "#C0392B"
col_fin_180 = "#2E8B5F" if pct_fin_180 >= 25 else "#D97706" if pct_fin_180 >= 10 else "#C0392B"

# ─── CHALLENGE ────────────────────────────────────────────────────────
_ch = _xl["🏁 Challenge"]
_ch_r = [_ch.iloc[i].to_list() for i in range(25)]
ch_periodo = s(_ch_r[2][1]) if _ch_r[2][1] else "N/D"
ch_min_pol = ni(_ch_r[2][6])
ch_min_pa  = n(_ch_r[3][6])

ch_partecipanti = []
for _i in range(6, 25):
    try:
        _nm = s(_ch_r[_i][6])
        if _nm and _nm != "nan" and isFB(_nm):
            ch_partecipanti.append(_nm)
    except:
        pass

ch_classifica = []
for _i in range(5, 22):
    try:
        _r2 = _ch_r[_i]
        _fb = s(_r2[1])
        if _fb and isFB(_fb):
            ch_classifica.append({"pos": s(_r2[0]), "fb": _fb, "pol": ni(_r2[2]), "pa": n(_r2[3])})
    except:
        pass

_medals_ch = ["&#x1F947;","&#x1F948;","&#x1F949;"] + [f"{_i2}&#xB0;" for _i2 in range(4,20)]
ch_rows = ""
for _i, _row in enumerate(ch_classifica):
    _vp = ch_min_pol > 0 and _row["pol"] >= ch_min_pol
    _va = ch_min_pa  > 0 and _row["pa"]  >= ch_min_pa
    _vince = _vp and (ch_min_pa == 0 or _va)
    _bg = "background:linear-gradient(90deg,rgba(200,169,81,.1),transparent)" if _vince else ""
    _bp = tag("tg", f"&#x2713; {_row['pol']} pol.") if _vp else tag("tr2" if _row["pol"]==0 else "ta", f"{_row['pol']} pol.")
    _ba = tag("tg", f"&#x2713; {fe(_row['pa'])}") if _va else tag("tr2" if _row["pa"]==0 else "ta", fe(_row["pa"]))
    _stato = "&#x1F3C6; Vincitore" if _vince else ("&#x23F3; In corsa" if _row["pol"]>0 else "&#x274C; Nessuna pol.")
    ch_rows += (
        f"<tr style='{_bg}'>"
        f"<td>{_medals_ch[_i] if _i<len(_medals_ch) else _i+1}</td>"
        f"<td><strong>{_row['fb']}</strong></td>"
        f"<td>{_bp}</td><td>{_ba}</td>"
        f"<td style='font-size:.8rem'>{_stato}</td></tr>"
    )

n_partecipanti = len(ch_partecipanti)
ch_pills = "".join(
    f"<span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>{_nm}</span>"
    for _nm in ch_partecipanti
)

# Variabili per form interattivo
import datetime as _dt3
def _fmt_date(v):
    if hasattr(v, "strftime"): return v.strftime("%Y-%m-%d")
    try:
        s2 = str(v)[:10]
        return s2 if len(s2)==10 else ""
    except: return ""

ch_inizio_val = _fmt_date(_ch_r[0][6])
ch_fine_val   = _fmt_date(_ch_r[1][6])
ch_min_pa_val = int(ch_min_pa) if ch_min_pa > 0 else 0

# Checkboxes per tutti i FB
ch_checkboxes = ""
for _c2 in D["collaboratori"]:
    _nm2 = _c2["name"]
    _checked = "checked" if _nm2 in ch_partecipanti else ""
    ch_checkboxes += (
        f"<label style='display:inline-flex;align-items:center;gap:5px;background:{'var(--cream)' if not _checked else 'rgba(200,169,81,.15)'};border:1px solid {'var(--brd)' if not _checked else 'var(--gold)'};border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'>"
        f"<input type='checkbox' {'checked' if _checked else ''} value='{_nm2}' onchange='updateChallenge()' style='accent-color:var(--gold)'>"
        f"{_nm2}</label>"
    )
ch_badge_pol = (tag("bb", f"Min. Polizze: {ch_min_pol}") if ch_min_pol > 0 else "")
ch_badge_pa  = (tag("bn", f"Min. Premio: {fe(ch_min_pa)}") if ch_min_pa > 0 else "")
ch_obj_details = ""
if ch_min_pol > 0:
    ch_obj_details += f"<p style='font-size:.8rem;margin-bottom:6px'>&#x1F4CB; <strong>Polizze minime:</strong> {ch_min_pol}</p>"
if ch_min_pa > 0:
    ch_obj_details += f"<p style='font-size:.8rem'>&#x1F4B6; <strong>Premio annuo minimo:</strong> {fe(ch_min_pa)}</p>"
if not ch_obj_details:
    ch_obj_details = "<p style='font-size:.8rem;color:#64748B'>Nessun obiettivo impostato</p>"

# JSON per JS interattivo
import json as _json2
ch_dati_json  = _json2.dumps(ch_classifica, ensure_ascii=True)
ch_originali_json = _json2.dumps(ch_partecipanti, ensure_ascii=True)
ch_min_pa_val_js = int(ch_min_pa) if ch_min_pa > 0 else 0

# ─── CHALLENGE LONDRA — calcola direttamente da Controllo Rate ───────
# NON legge il foglio Challenge Londra (valori cached), ma ricalcola
# da zero leggendo Controllo Rate con openpyxl (data_only=True).
import datetime as _dt2
from openpyxl import load_workbook as _lw

_LON_START   = _dt2.datetime(2026, 5, 1)
_LON_END     = _dt2.datetime(2026, 6, 30)
_LON_MIN_POL = 4
_LON_MIN_INC = 5000.0

# Leggi i parametri dal foglio Challenge Londra (celle fisse, non formule)
try:
    _chl_cfg = xl['🏁 Challenge Londra']
    _p = _chl_cfg.iloc[2].to_list()
    _q = _chl_cfg.iloc[3].to_list()
    if ni(_p[6]) > 0: _LON_MIN_POL = ni(_p[6])
    if n(_q[6])  > 0: _LON_MIN_INC = n(_q[6])
except: pass

# Partecipanti dal foglio Challenge Londra
_lon_all_fb = []
try:
    _chl_r = [_chl_cfg.iloc[i].to_list() for i in range(26)]
    for _i in range(5, 26):
        _nm = s(_chl_r[_i][6])
        if _nm and _nm != 'nan' and isFB(_nm):
            _lon_all_fb.append(_nm)
except: pass
if not _lon_all_fb:
    _lon_all_fb = [c['name'] for c in D['collaboratori'] if c.get('gruppo') == 'Onorato']

# Calcola incassato direttamente da Controllo Rate con openpyxl
_lon_inc = {}
_lon_pol = {}
try:
    _wb2 = _lw(EXCEL, data_only=True)
    _ws2 = _wb2['💳 Controllo Rate']
    for _rn in range(5, 1004):
        _fb2 = _ws2.cell(_rn, 2).value
        if not _fb2 or str(_fb2).strip() in ('nan', ''): break
        _dp2 = _ws2.cell(_rn, 7).value
        if not isinstance(_dp2, _dt2.datetime): continue
        if not (_LON_START <= _dp2 <= _LON_END): continue
        _fb2  = str(_fb2).strip()
        _pf2  = float(_ws2.cell(_rn, 5).value or 0)
        _pa2  = float(_ws2.cell(_rn, 8).value or 0)
        _tip2 = str(_ws2.cell(_rn, 4).value or '').strip()
        _mag2 = _ws2.cell(_rn, 14).value
        _giu2 = _ws2.cell(_rn, 15).value
        # Escludi: Protezione Assistenza sempre
        # Escludi: Protezione Casa e Famiglia con premio ANNUO <= 250€
        if _tip2 == 'Protezione Assistenza': continue
        if _tip2 == 'Protezione Casa e Famiglia' and _pa2 <= 250: continue
        # Conta solo se ha almeno una rata incassata (✓) in maggio o giugno
        if _mag2 != '✓' and _giu2 != '✓': continue
        _lon_pol[_fb2] = _lon_pol.get(_fb2, 0) + 1
        if _mag2 == '✓':
            _lon_inc[_fb2] = _lon_inc.get(_fb2, 0.0) + _pf2
        if _giu2 == '✓':
            _lon_inc[_fb2] = _lon_inc.get(_fb2, 0.0) + _pf2
    _wb2.close()
except Exception as _e:
    print(f"Warning calcolo Londra: {_e}")

# Costruisci classifica con tutti i partecipanti
_lon_all_fb = [fb for fb in _lon_all_fb if fb != 'Maola Daniele']

_lon_class = []
for _fb3 in _lon_all_fb:
    _lon_class.append({
        'fb':  _fb3,
        'pol': _lon_pol.get(_fb3, 0),
        'inc': round(_lon_inc.get(_fb3, 0.0), 2)
    })
# Aggiungi eventuali FB con polizze non nella lista partecipanti
for _fb3, _pv in _lon_pol.items():
    if _fb3 == 'Maola Daniele': continue
    if not any(r['fb'] == _fb3 for r in _lon_class):
        _lon_class.append({'fb': _fb3, 'pol': _pv, 'inc': round(_lon_inc.get(_fb3, 0.0), 2)})

_lon_class.sort(key=lambda x: (-x['inc'], -x['pol'], x['fb']))

_lon_medals   = ['&#x1F947;','&#x1F948;','&#x1F949;'] + [f'{_i2}&#xB0;' for _i2 in range(4, 30)]
_lon_oggi     = _dt2.datetime.today().strftime('%d/%m/%Y')
_lon_in_corso = _dt2.datetime.today() >= _LON_START
lon_stato_bar = (
    f'<span class="tag tg" style="margin-left:auto">&#x1F7E2; Challenge in corso &mdash; agg. {_lon_oggi}</span>'
    if _lon_in_corso else '<span class="tag ta">&#x26A0; Challenge non ancora iniziata</span>'
)

lon_rows = ""
_pos3 = 0
for _row3 in _lon_class:
    if _row3['fb'] == 'Maola Daniele': continue
    _vp3  = _row3['pol'] >= _LON_MIN_POL
    _vi3  = _row3['inc'] >= _LON_MIN_INC
    _win3 = _vp3 and _vi3
    _bg3  = "background:linear-gradient(90deg,rgba(200,169,81,.1),transparent)" if _win3 else ""
    _bp3  = (f"<span class='tag tg'>&#x2713; {_row3['pol']} pol.</span>" if _vp3
             else f"<span class='tag {'tr2' if _row3['pol']==0 else 'ta'}'>{_row3['pol']} pol.</span>")
    _bi3  = (f"<span class='tag tg'>&#x2713; {fe(_row3['inc'])}</span>" if _vi3
             else f"<span class='tag {'tr2' if _row3['inc']==0 else 'ta'}'>{fe(_row3['inc'])}</span>")
    _gap3 = _LON_MIN_INC - _row3['inc']
    _bgap3 = "<span class='tag tg'>&#x2713; Raggiunto</span>" if _gap3 <= 0 else f"<span class='tag ta'>- {fe(_gap3)}</span>"
    _st3  = ("&#x1F3C6; Vincitore" if _win3
             else "&#x23F3; Serve inc." if _vp3
             else "&#x23F3; In corsa" if _row3['pol'] > 0
             else "&#x274C; No pol.")
    _med3 = _lon_medals[_pos3] if _pos3 < len(_lon_medals) else str(_pos3 + 1)
    lon_rows += (f"<tr style='{_bg3}'><td>{_med3}</td><td><strong>{_row3['fb']}</strong></td>"
                 f"<td>{_bp3}</td><td>{_bi3}</td><td>{_bgap3}</td><td style='font-size:.75rem'>{_st3}</td></tr>")
    _pos3 += 1

lon_pills  = "".join(
    f"<span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>{_fb3}</span>"
    for _fb3 in _lon_all_fb)
lon_n_part = len(_lon_all_fb)

# ─── POLIZZE IN LAVORAZIONE ─────────────────────────────────────────
lav_list = sorted(D['polLavorazione'], key=lambda x: (x['fb'], x['cliente']))
lav_rows = ""
for _l in lav_list:
    lav_rows += (f"<tr><td>{_l['fb']}</td><td>{_l['cliente']}</td>"
                 f"<td>{_l['tipoPol']}</td>"
                 f"<td class='num'>{fe(_l['premioFirma'])}</td>"
                 f"<td class='num'>{fe(_l['premioAnnuo'])}</td></tr>")
n_lav = len(lav_list)
tot_lav_pa = fe(sum(_l['premioAnnuo'] for _l in lav_list))

# ─── PORTAFOGLIO CLIENTI ─────────────────────────────────────────────
# Build from Dati Giornalieri + Controllo Rate
# cr_lookup: (fb, cliente, tipoPol) -> {data_polizza, stato, premioFirma}
_cr_lookup = {}
try:
    _ct_pf = xl['💳 Controllo Rate']
    import datetime as _dtt
    for _ii in range(3, 600):
        try: _rr = _ct_pf.iloc[_ii].to_list()
        except: break
        _fb_cr = s(_rr[1]); _cli_cr = s(_rr[2]); _tip_cr = s(_rr[3])
        if not _fb_cr or not _cli_cr: continue
        _dp_cr = _rr[6]
        if hasattr(_dp_cr, 'to_pydatetime'): _dp_cr = _dp_cr.to_pydatetime()
        _key = (_fb_cr, _cli_cr, _tip_cr)
        _cr_lookup[_key] = {
            'data_polizza': _dp_cr if isinstance(_dp_cr, _dtt.datetime) else None,
            'stato': s(_rr[8]),
            'premioFirma': n(_rr[4]),
            'premioAnnuo': n(_rr[7])
        }
except Exception as _e:
    pass

pf_rows = ""
_pf_count = {'proc': 0, 'lav': 0, 'ann': 0}
for _r in sorted(D['giornalieri'], key=lambda x: (x.get('data') if hasattr(x.get('data'), 'strftime') else __import__('datetime').datetime.min), reverse=True):
    if _r['esito'] != 'Sottoscritto': continue
    _fb2 = _r['fb']; _cli2 = _r['cliente']; _tip2 = _r['tipoPol']
    _pf2 = _r['premioFirma']; _pa2 = _r['premioAnnuo']
    _fraz2 = _r['frazionamento']
    _da2 = _r.get('data')
    _da_str = _da2.strftime('%d/%m/%Y') if isinstance(_da2, __import__('datetime').datetime) else '—'
    # Get from CR
    _cr = _cr_lookup.get((_fb2, _cli2, _tip2), {})
    _dp_str = _cr['data_polizza'].strftime('%d/%m/%Y') if _cr.get('data_polizza') else '—'
    _stato = _cr.get('stato', '—')
    _pf_real = _cr.get('premioFirma', _pf2)
    # Status tag
    if _stato == 'Processata':
        _stag = "<span class='tag tg'>&#x2713; Processata</span>"
        _pf_count['proc'] += 1
    elif _stato == 'In lavorazione':
        _stag = "<span class='tag ta'>&#x23F3; In lavorazione</span>"
        _pf_count['lav'] += 1
    elif _stato == 'Annullata':
        _stag = "<span class='tag tr2'>&#x274C; Annullata</span>"
        _pf_count['ann'] += 1
    else:
        _stag = f"<span class='tag tn'>{_stato}</span>"
    pf_rows += (
        f"<tr>"
        f"<td>{_fb2}</td>"
        f"<td>{_cli2}</td>"
        f"<td style='font-size:.75rem'>{_tip2}</td>"
        f"<td style='font-size:.75rem;text-align:center'>{_fraz2}</td>"
        f"<td class='num' style='text-align:center'>{fe(_pf_real)}</td>"
        f"<td style='text-align:center;font-size:.75rem'>{_da_str}</td>"
        f"<td style='text-align:center;font-size:.75rem'>{_dp_str}</td>"
        f"<td style='text-align:center'>{_stag}</td>"
        f"</tr>"
    )
pf_proc = _pf_count['proc']
pf_lav  = _pf_count['lav']
pf_ann  = _pf_count['ann']

# Polizze mensili/annuali e premio medio (solo processate)
_pf_mensile  = sum(1 for _r in D['giornalieri'] if _r['esito'] == 'Sottoscritto' and _cr_lookup.get((_r['fb'], _r['cliente'], _r['tipoPol']), {}).get('stato') == 'Processata' and _r['frazionamento'] == 'Mensile')
_pf_annuale  = sum(1 for _r in D['giornalieri'] if _r['esito'] == 'Sottoscritto' and _cr_lookup.get((_r['fb'], _r['cliente'], _r['tipoPol']), {}).get('stato') == 'Processata' and _r['frazionamento'] == 'Annuale')
_pf_tot_pf   = sum(_r['premioFirma'] for _r in D['giornalieri'] if _r['esito'] == 'Sottoscritto' and _cr_lookup.get((_r['fb'], _r['cliente'], _r['tipoPol']), {}).get('stato') == 'Processata')
_pf_med      = round(_pf_tot_pf / max(pf_proc, 1))
pf_mensile   = _pf_mensile
pf_annuale   = _pf_annuale
pf_med       = fe(_pf_med)

CSS = """
:root{--navy:#0B1E3D;--n2:#142952;--n3:#1E3A6E;--gold:#C8A951;--g2:#E8CC7A;--cream:#FAF6EE;--w:#fff;--gr:#2E8B5F;--gr2:#3BA870;--red:#C0392B;--amb:#D97706;--mut:#64748B;--brd:rgba(200,169,81,.2);--sh:0 2px 14px rgba(11,30,61,.07);--sh2:0 6px 28px rgba(11,30,61,.13)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'DM Sans',sans-serif;background:var(--cream);color:var(--navy)}
.num{font-family:'Outfit',sans-serif;font-weight:600}
.hdr{background:linear-gradient(135deg,var(--navy),var(--n3));padding:14px 32px;display:flex;justify-content:space-between;align-items:center;border-bottom:2px solid var(--gold);position:sticky;top:0;z-index:100;box-shadow:0 2px 20px rgba(0,0,0,.25);gap:16px}
.hdr h1{font-family:'Playfair Display',serif;color:var(--g2);font-size:1.45rem;font-weight:700;line-height:1}
.hdr p{color:rgba(255,255,255,.42);font-size:.68rem;text-transform:uppercase;letter-spacing:.07em;margin-top:2px}
.nav{display:flex;gap:4px;flex-wrap:wrap}
.nb{background:rgba(255,255,255,.07);border:1px solid rgba(255,255,255,.13);color:rgba(255,255,255,.68);padding:6px 14px;border-radius:4px;cursor:pointer;font-size:.74rem;font-family:'DM Sans',sans-serif;transition:all .18s}
.nb.on,.nb:hover{background:var(--gold);color:var(--navy);border-color:var(--gold);font-weight:600}
.wrap{max-width:1520px;margin:0 auto;padding:26px 36px}
.sec{display:none}.sec.on{display:block}
.st{font-family:'Playfair Display',serif;font-size:1.2rem;color:var(--navy);font-weight:600;margin-bottom:3px}
.ss{color:var(--mut);font-size:.68rem;text-transform:uppercase;letter-spacing:.06em;margin-bottom:20px}
.g5{display:grid;grid-template-columns:repeat(5,1fr);gap:13px;margin-bottom:16px}
.g4{display:grid;grid-template-columns:repeat(4,1fr);gap:13px;margin-bottom:16px}
.g3{display:grid;grid-template-columns:repeat(3,1fr);gap:13px;margin-bottom:16px}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}
.hero{background:linear-gradient(135deg,var(--navy),var(--n3));border-radius:12px;padding:24px 30px;margin-bottom:16px;box-shadow:var(--sh2);display:grid;grid-template-columns:1fr 110px;gap:18px;align-items:center;position:relative;overflow:hidden}
.hero::after{content:'';position:absolute;right:-35px;top:-35px;width:160px;height:160px;border:1px solid rgba(200,169,81,.15);border-radius:50%}
.hl{font-size:.62rem;text-transform:uppercase;letter-spacing:.1em;color:var(--g2);margin-bottom:5px}
.ht{font-family:'Playfair Display',serif;font-size:1.5rem;color:#fff;margin-bottom:4px}
.hs{color:rgba(255,255,255,.48);font-size:.76rem}
.hp{background:rgba(255,255,255,.1);border-radius:4px;height:6px;overflow:hidden;margin-top:10px}
.hpb{height:100%;border-radius:4px;background:linear-gradient(90deg,var(--gold),var(--g2))}
.hf{color:rgba(255,255,255,.3);font-size:.63rem;margin-top:4px}
.hpct{font-family:'Outfit',sans-serif;font-size:3rem;color:var(--g2);font-weight:700;line-height:1;text-align:right}
.hpcts{color:rgba(255,255,255,.38);font-size:.67rem;text-align:right;margin-top:2px}
.card{background:var(--w);border-radius:10px;padding:18px 20px;box-shadow:var(--sh);border-top:3px solid var(--n3);position:relative;transition:transform .18s,box-shadow .18s}
.card:hover{transform:translateY(-2px);box-shadow:var(--sh2)}
.card.gold{border-top-color:var(--gold)}.card.green{border-top-color:var(--gr)}.card.red{border-top-color:var(--red)}.card.amb{border-top-color:var(--amb)}.card.navy{border-top-color:var(--n2)}
.cico{position:absolute;right:14px;top:14px;width:32px;height:32px;background:var(--cream);border-radius:7px;display:flex;align-items:center;justify-content:center;font-size:.95rem}
.cl{font-size:.6rem;text-transform:uppercase;letter-spacing:.08em;color:var(--mut);font-weight:500;margin-bottom:6px}
.cv{font-family:'Outfit',sans-serif;font-size:1.75rem;color:var(--navy);font-weight:700;line-height:1;margin-bottom:3px}
.csub{font-size:.7rem;color:var(--mut);line-height:1.4}
.bdg{display:inline-block;padding:2px 8px;border-radius:10px;font-size:.62rem;font-weight:600;margin-top:5px}
.bg{background:#D1FAE5;color:#065F46}.br{background:#FEE2E2;color:#991B1B}.ba{background:#FEF3C7;color:#92400E}.bb{background:#DBEAFE;color:#1E40AF}.bn{background:#EEF2FF;color:var(--navy)}
.prog{background:var(--cream);border-radius:3px;height:5px;margin-top:7px;overflow:hidden}
.pb{height:100%;border-radius:3px}
.pg{background:linear-gradient(90deg,var(--gr),var(--gr2))}.pa{background:linear-gradient(90deg,var(--amb),#F59E0B)}.pr{background:linear-gradient(90deg,var(--red),#E74C3C)}
.onc{background:var(--w);border-radius:12px;box-shadow:var(--sh);margin-bottom:16px;border-top:3px solid var(--gold);overflow:hidden}
.onh{padding:13px 20px;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid var(--brd)}
.onh h3{font-family:'Playfair Display',serif;font-size:.9rem;color:var(--navy);font-weight:600}
.onh span{font-size:.62rem;color:var(--mut);text-transform:uppercase;letter-spacing:.05em}
.onb{padding:18px 20px;display:grid;grid-template-columns:repeat(5,1fr) 1.6fr;gap:14px;align-items:center}
.ohl{font-size:.58rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:3px}
.ohv{font-family:'Outfit',sans-serif;font-size:1.25rem;font-weight:700;line-height:1}
.ohs{font-size:.62rem;color:var(--mut);margin-top:2px}
.tw{background:var(--w);border-radius:12px;overflow:hidden;box-shadow:var(--sh);margin-bottom:16px}
.twh{padding:12px 18px;border-bottom:1px solid var(--brd);display:flex;justify-content:space-between;align-items:center}
.twh h3{font-family:'Playfair Display',serif;font-size:.87rem;color:var(--navy);font-weight:600}
.twh span{font-size:.62rem;color:var(--mut);text-transform:uppercase;letter-spacing:.04em}
table{width:100%;border-collapse:collapse}
thead th{background:var(--navy);color:rgba(255,255,255,.8);font-size:.58rem;text-transform:uppercase;letter-spacing:.07em;padding:11px 11px;text-align:center;font-weight:500;white-space:normal;line-height:1.4;vertical-align:middle}
thead th:first-child{padding-left:18px}
tbody tr{border-bottom:1px solid rgba(11,30,61,.04)}
tbody tr:hover{background:rgba(11,30,61,.016)}
tbody td{padding:9px 11px;font-size:.77rem}
tbody td:first-child{padding-left:18px;font-weight:500}
.tag{display:inline-block;padding:2px 7px;border-radius:9px;font-size:.62rem;font-weight:600}
.tg{background:#D1FAE5;color:#065F46}.tr2{background:#FEE2E2;color:#991B1B}.ta{background:#FEF3C7;color:#92400E}.tb{background:#DBEAFE;color:#1E40AF}.tn{background:#EEF2FF;color:var(--navy)}
.rkc{background:var(--w);border-radius:12px;overflow:hidden;box-shadow:var(--sh)}
.rkh{padding:11px 15px;background:var(--navy);color:#fff;font-family:'Playfair Display',serif;font-size:.82rem;font-weight:600}
.rki{display:flex;align-items:center;padding:9px 15px;border-bottom:1px solid rgba(11,30,61,.05);gap:8px}
.rki:last-child{border-bottom:none}
.rip{width:24px;text-align:center;font-size:.85rem;flex-shrink:0}
.rin{flex:1;font-size:.77rem;font-weight:500}
.riv{font-family:'Outfit',sans-serif;font-size:.77rem;font-weight:700;color:var(--navy);text-align:right}
.ris{font-size:.62rem;color:var(--mut);text-align:right}
.bch{background:var(--w);border-radius:12px;padding:18px;box-shadow:var(--sh);margin-bottom:16px}
.bcht{font-family:'Playfair Display',serif;font-size:.87rem;color:var(--navy);margin-bottom:14px;font-weight:600}
.brr{display:flex;align-items:center;gap:8px;margin-bottom:8px}
.brl{width:155px;font-size:.71rem;flex-shrink:0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.bro{flex:1;background:var(--cream);border-radius:3px;height:18px;overflow:hidden}
.bri{height:100%;border-radius:3px;display:flex;align-items:center;padding-left:6px;min-width:2px}
.bri span{font-size:.62rem;font-weight:600;color:#fff;white-space:nowrap}
.brn{width:125px;text-align:right;font-size:.69rem;color:var(--mut);flex-shrink:0;white-space:nowrap}
.alerts{display:grid;grid-template-columns:1fr 1fr;gap:13px;margin-bottom:16px}
.alert{background:var(--w);border-radius:10px;padding:16px 18px;box-shadow:var(--sh);border-left:4px solid}
.alert.d{border-color:var(--red)}.alert.w{border-color:var(--amb)}.alert.ok{border-color:var(--gr)}.alert.i{border-color:var(--n3)}
.alt{font-size:.62rem;text-transform:uppercase;letter-spacing:.08em;font-weight:700;margin-bottom:6px}
.alert.d .alt{color:var(--red)}.alert.w .alt{color:var(--amb)}.alert.ok .alt{color:var(--gr)}.alert.i .alt{color:var(--n3)}
.alb{font-size:.77rem;line-height:1.55}
.alb ul{margin-left:12px;margin-top:3px}
.alb li{margin-bottom:2px}
.trw{background:var(--w);border-radius:12px;padding:18px;box-shadow:var(--sh);margin-bottom:16px;overflow-x:auto}
.trt{font-family:'Playfair Display',serif;font-size:.87rem;color:var(--navy);margin-bottom:14px;font-weight:600}
.tt{border-collapse:collapse;width:100%;min-width:550px}
.tt th{background:var(--navy);color:rgba(255,255,255,.8);font-size:.58rem;text-transform:uppercase;letter-spacing:.06em;padding:7px 10px;text-align:center;font-weight:500}
.tt th:first-child{text-align:left;padding-left:14px}
.tt td{padding:7px 10px;font-size:.74rem;text-align:center;border-bottom:1px solid rgba(11,30,61,.038)}
.tt td:first-child{text-align:left;padding-left:14px;font-weight:500}
.tt .totr td{background:rgba(11,30,61,.05);font-weight:600}
.ck{display:block;text-align:center;padding:2px 4px;border-radius:3px;font-size:.68rem;font-family:'Outfit',sans-serif;font-weight:500}
.ckok{background:#D1FAE5;color:#065F46;font-weight:700}.ckw{background:#FEF3C7;color:#92400E;font-weight:700}.ckr{background:#FEE2E2;color:#991B1B;font-weight:700}.ck0{color:var(--mut)}
.chw{background:var(--w);border-radius:12px;padding:18px;box-shadow:var(--sh);margin-bottom:16px}
.chw h3{font-family:'Playfair Display',serif;font-size:.87rem;color:var(--navy);font-weight:600;margin-bottom:14px}
.coll{background:var(--w);border-radius:12px;box-shadow:var(--sh);overflow:hidden;margin-bottom:16px}
.collh{background:linear-gradient(135deg,var(--navy),var(--n3));padding:14px 22px;display:flex;align-items:center;gap:11px;border-bottom:2px solid var(--gold)}
.collh h3{font-family:'Playfair Display',serif;color:var(--g2);font-size:.92rem;font-weight:600;flex:1}
.collsel{font-family:'DM Sans',sans-serif;font-size:.77rem;padding:6px 11px;border:1px solid rgba(200,169,81,.4);border-radius:5px;background:rgba(255,255,255,.08);color:#fff;cursor:pointer;min-width:195px;outline:none}
.collsel option{color:var(--navy);background:#fff}
.collb{padding:22px}
.collph{text-align:center;padding:28px;color:var(--mut);font-size:.82rem}
.collg{display:grid;grid-template-columns:230px 1fr;gap:20px;align-items:start}
.cpf{background:var(--cream);border-radius:10px;padding:18px;text-align:center;min-width:0;overflow:hidden}
.cpav{width:56px;height:56px;background:linear-gradient(135deg,var(--navy),var(--n3));border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:1.3rem;margin:0 auto 9px;color:var(--g2);font-family:'Playfair Display',serif;font-weight:700}
.cpnm{font-family:'Playfair Display',serif;font-size:.95rem;color:var(--navy);font-weight:600;margin-bottom:2px}
.cpfb{font-size:.62rem;color:var(--mut);text-transform:uppercase;letter-spacing:.05em;margin-bottom:10px}
.cpst{display:flex;justify-content:space-between;padding:5px 0;border-bottom:1px solid rgba(11,30,61,.05);font-size:.73rem}
.cpst:last-child{border-bottom:none}
.csl{color:var(--mut)}.csv{font-weight:600;color:var(--navy);font-family:'Outfit',sans-serif}
.ckg{display:grid;grid-template-columns:repeat(3,1fr);gap:9px;margin-bottom:14px}
.ckc{background:var(--cream);border-radius:7px;padding:11px 13px;border-top:3px solid var(--n3)}
.ckc.gold{border-top-color:var(--gold)}.ckc.green{border-top-color:var(--gr)}.ckc.red{border-top-color:var(--red)}.ckc.amb{border-top-color:var(--amb)}
.ckl{font-size:.57rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:3px}
.ckv{font-family:'Outfit',sans-serif;font-size:1.15rem;color:var(--navy);font-weight:700}
.cks{font-size:.62rem;color:var(--mut);margin-top:2px}
.csec{margin-bottom:14px}
.csect{font-family:'Playfair Display',serif;font-size:.82rem;color:var(--navy);font-weight:600;margin-bottom:8px;padding-bottom:4px;border-bottom:1px solid var(--brd)}
.mt{width:100%;border-collapse:collapse;font-size:.72rem}
.mt th{background:rgba(11,30,61,.05);color:var(--navy);font-size:.57rem;text-transform:uppercase;letter-spacing:.06em;padding:6px 8px;text-align:left;font-weight:600}
.mt td{padding:6px 8px;border-bottom:1px solid rgba(11,30,61,.04)}
.mt tr:hover td{background:rgba(11,30,61,.012)}
.anl{background:var(--cream);border-radius:8px;padding:13px 15px}
.anl h4{font-family:'Playfair Display',serif;font-size:.82rem;color:var(--navy);margin-bottom:9px;font-weight:600}
.anlr{display:flex;gap:7px;margin-bottom:7px;align-items:flex-start}
.anlt{font-size:.75rem;line-height:1.5}
@media(max-width:1100px){.g5,.g4{grid-template-columns:repeat(2,1fr)}.onb{grid-template-columns:repeat(3,1fr)}}
@media(max-width:700px){.g5,.g4,.g3,.g2,.alerts{grid-template-columns:1fr}.collg{grid-template-columns:1fr}.wrap{padding:14px}.hdr{flex-direction:column;gap:8px;padding:12px}}
"""

# ─── ASSEMBLE HTML ──────────────────────────────────────────────────
HTML = f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>FPS Dashboard 2026</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=DM+Sans:wght@300;400;500;600&family=Outfit:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>{CSS}</style>
</head>
<body>
<header class="hdr" style="gap:16px">
  <div style="flex-shrink:0;line-height:1.2">
    <div style="font-family:'Playfair Display',serif;color:var(--g2);font-size:1.25rem;font-weight:600;letter-spacing:.01em">&#x1F6E1;&#xFE0F; Dashboard</div>
    <div style="font-family:'Playfair Display',serif;color:var(--g2);font-size:1.25rem;font-weight:400;font-style:italic;letter-spacing:.02em">FPS Maola Daniele</div>
    <div style="font-family:'Playfair Display',serif;color:var(--g2);font-size:.82rem;font-weight:400;font-style:italic;letter-spacing:.02em;margin-top:2px;opacity:.85">&ldquo;Non c&rsquo;&egrave; Pianificazione senza Protezione&rdquo;</div>
  </div>
  <nav class="nav" style="flex:1;justify-content:flex-end;align-items:center">
    <button class="nb on" onclick="showSec('ov')">KPI Generale</button>
    <button class="nb" onclick="showSec('fb')">KPI Family Banker</button>
    <button class="nb" onclick="showSec('tr')">Andamento Mensile</button>
    <button class="nb" onclick="showSec('obj')">&#x1F4C8; Obiettivi</button>
    <button class="nb" onclick="showSec('ch')">&#x1F3C6; Challenge</button>
    <button class="nb" onclick="showSec('az')">Azioni Correttive</button>
    <button class="nb" onclick="showSec('pf')">&#x1F4C1; Portafoglio</button>
    <span style="color:rgba(255,255,255,.35);font-size:.62rem;letter-spacing:.05em;white-space:nowrap;margin-left:4px">Agg. {oggi}</span>
  </nav>
</header>
<div class="wrap">

<section class="sec on" id="s-ov">
  <p class="st">KPI Generale &mdash; YTD 2026</p>
  <p class="ss">Panoramica complessiva &middot; {oggi}</p>
  <div class="hero">
    <div>
      <p class="hl">Obiettivo annuo premi incassati</p>
      <p class="ht">{fe(G['target'])} &mdash; Target 2026</p>
      <p class="hs">Incassato ad oggi: <strong style="color:var(--g2)">{fe(G['premiIncassati'])}</strong> &nbsp;|&nbsp; Gap: <strong style="color:var(--g2)">{fe(G['target']-G['premiIncassati'])}</strong></p>
      <div class="hp"><div class="hpb" style="width:{pct_target}%"></div></div>
      <p class="hf">{pct_target}% raggiunto &mdash; 253 giorni residui all'anno</p>
    </div>
    <div><p class="hpct">{pct_target}%</p><p class="hpcts">del budget</p></div>
  </div>
  <div class="g2">{kpi1}</div>
  <div class="g5">{kpi2}</div>
  <div class="g4">{kpi3}</div>
  <div class="onc"><div class="onh"><h3>&#x1F3AF; Obiettivo Gruppo Onorato 2026</h3><span>Fonte: Foglio Obj Onorato</span></div><div class="onb">{onb}</div></div>
  <div class="g2">
    <div class="tw">
      <div class="twh"><h3>&#x1F3C6; Ranking &mdash; Premio Annuo YTD</h3><span>Top 10</span></div>
      <table><thead><tr><th>#</th><th>Family Banker</th><th>Appt</th><th>Pol.</th><th>Premio Annuo</th><th>Inc.</th><th>Conv.%</th></tr></thead>
      <tbody>{rank_rows}</tbody></table>
    </div>
    <div><div class="bch">{prod_html}</div><div class="g3">{top3}</div></div>
  </div>

  <!-- POLIZZE IN LAVORAZIONE -->
  <div class="tw" style="margin-top:20px">
    <div class="twh"><h3>&#x2699;&#xFE0F; Polizze in Lavorazione</h3><span>{n_lav} polizze &middot; Premio Annuo Totale: {tot_lav_pa}</span></div>
    <table><thead><tr><th>Family Banker</th><th>Cliente</th><th>Tipo Polizza</th><th>Premio Firma</th><th>Premio Annuo</th></tr></thead>
    <tbody>{lav_rows}</tbody></table>
  </div>

</section>

<section class="sec" id="s-fb">
  <p class="st">KPI per Family Banker</p>
  <p class="ss">Dettaglio individuale &middot; Schede colloquio</p>
  <div class="coll">
    <div class="collh">
      <h3>&#x1F464; Scheda Colloquio &mdash; Seleziona Collaboratore</h3>
      <select class="collsel" id="fbsel" onchange="showColl(this.value)">
        <option value="">&#x2014; Scegli il collaboratore &#x2014;</option>
        {fb_options}
      </select>
    </div>
    <div class="collb" id="collb"><div class="collph">Seleziona un Family Banker per la scheda completa del colloquio</div></div>
  </div>
  <div class="g4">{fbsum}</div>
  <div class="tw">
    <div class="twh"><h3>&#x1F465; Tutti i Family Banker &mdash; Performance YTD</h3><span>&#x25CF; Verde pol&ge;{MESE_CORR} &middot; &#x25CF; Arancione pol&ge;{MESE_CORR-2} &middot; &#x25CF; Rosso pol&lt;{MESE_CORR-2}</span></div>
    <div style="overflow-x:auto">
    <table><thead><tr><th>Family Banker</th><th>FBO</th><th>N°<br>App.ti</th><th>N°<br>Polizze</th><th>Attivo</th><th>Premio<br>Annuo</th><th>Premi<br>Incassati</th><th style="max-width:60px">Polizze in<br>Lavorazione</th><th style="max-width:80px">Premio alla Firma<br>in Lavorazione</th><th>CB</th></tr></thead>
    <tbody>{fb_rows}</tbody></table></div>
  </div>
  <div class="g3">{top5}</div>

  <!-- ANALISI PRODUZIONE PERSONALIZZATA -->
  <div class="tw" id="prod-widget" style="margin-top:20px">
    <div class="twh"><h3>&#x1F50D; Analisi Produzione Personalizzata</h3><span>Seleziona FB, periodo e metriche</span></div>
    <div style="padding:18px 22px">
      <div style="display:flex;gap:16px;flex-wrap:wrap;margin-bottom:16px;align-items:flex-end">
        <div style="flex:2;min-width:220px">
          <p style="font-size:.63rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:6px">Family Banker (uno o pi&#xF9;)</p>
          <div id="prod-fb-pills" style="display:flex;flex-wrap:wrap;gap:6px;padding:10px;background:var(--cream);border:1px solid var(--brd);border-radius:7px;min-height:42px;cursor:pointer" onclick="toggleProdFbDropdown()">
            <span id="prod-fb-placeholder" style="color:var(--mut);font-size:.8rem;align-self:center">Clicca per selezionare...</span>
          </div>
          <div id="prod-fb-dropdown" style="display:none;position:absolute;z-index:200;background:var(--w);border:1px solid var(--brd);border-radius:8px;box-shadow:0 8px 30px rgba(0,0,0,.12);padding:8px;max-height:260px;overflow-y:auto;width:300px">
            <div style="display:flex;gap:6px;padding:4px 0 8px">
              <button onclick="prodSelectAll()" style="font-size:.7rem;padding:3px 8px;border:1px solid var(--brd);border-radius:4px;background:var(--cream);cursor:pointer">Tutti</button>
              <button onclick="prodSelectNone()" style="font-size:.7rem;padding:3px 8px;border:1px solid var(--brd);border-radius:4px;background:var(--cream);cursor:pointer">Nessuno</button>
            </div>
            <div id="prod-fb-checkboxes"></div>
          </div>
        </div>
        <div style="flex:1;min-width:130px">
          <p style="font-size:.63rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:6px">Dal</p>
          <input type="date" id="prod-dal" style="width:100%;padding:8px 10px;border:1px solid var(--brd);border-radius:6px;font-family:'DM Sans',sans-serif;font-size:.82rem;background:var(--cream);color:var(--navy)">
        </div>
        <div style="flex:1;min-width:130px">
          <p style="font-size:.63rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:6px">Al</p>
          <input type="date" id="prod-al" style="width:100%;padding:8px 10px;border:1px solid var(--brd);border-radius:6px;font-family:'DM Sans',sans-serif;font-size:.82rem;background:var(--cream);color:var(--navy)">
        </div>
        <div>
          <button onclick="runProdAnalysis()" style="padding:9px 22px;background:var(--navy);color:#fff;border:none;border-radius:7px;font-family:'DM Sans',sans-serif;font-size:.82rem;font-weight:600;cursor:pointer;white-space:nowrap">&#x1F50D; Analizza</button>
        </div>
      </div>
      <div style="margin-bottom:16px">
        <p style="font-size:.63rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:8px">Metriche da mostrare</p>
        <div style="display:flex;flex-wrap:wrap;gap:8px" id="prod-metrics">
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="pa" checked> Premio Annuo</label>
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="inc" checked> Incassato</label>
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="pol" checked> N&#xB0; Polizze</label>
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="appt" checked> N&#xB0; Appuntamenti</label>
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="cb"> N&#xB0; Callback</label>
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="lav"> Polizze in Lavorazione</label>
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="conv"> % Conversione</label>
          <label style="display:flex;align-items:center;gap:5px;padding:6px 12px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;cursor:pointer;font-size:.78rem"><input type="checkbox" value="med"> Premio Medio</label>
        </div>
      </div>
      <div id="prod-results" style="display:none">
        <div id="prod-kpi-cards" style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:16px"></div>
        <div style="overflow-x:auto"><table><thead id="prod-thead"></thead><tbody id="prod-tbody"></tbody></table></div>
        <p id="prod-note" style="font-size:.68rem;color:var(--mut);margin-top:8px"></p>
      </div>
      <div id="prod-empty" style="text-align:center;padding:24px;color:var(--mut);font-size:.82rem;font-style:italic">Seleziona i Family Banker, il periodo e le metriche, poi clicca Analizza.</div>
    </div>
  </div>

</section>

<section class="sec" id="s-tr">
  <p class="st">Andamento Mensile 2026</p>
  <p class="ss">Polizze, premi e incassati mese per mese</p>
  <div class="g4">{tsum}</div>
  <div class="g2">
    <div class="chw"><h3>&#x1F4B6; Premio alla Firma Mensile &mdash; Team Totale</h3>{SVG_PA}</div>
    <div class="chw"><h3>&#x2705; Premi Incassati Mensili &mdash; Team Totale</h3>{SVG_PI}</div>
  </div>
  <div class="trw"><p class="trt">&#x1F4CB; Polizze Sottoscritte per Family Banker per Mese</p>
    <table class="tt"><thead><tr><th>Family Banker</th>{hdr_mesi}<th>TOT</th></tr></thead>
    <tbody>{trend_pol_rows}</tbody></table></div>
  <div class="trw"><p class="trt">&#x1F4C5; Appuntamenti Effettuati per Family Banker per Mese</p>
    <table class="tt"><thead><tr><th>Family Banker</th>{hdr_mesi}<th>TOT</th></tr></thead>
    <tbody>{trend_appt_rows}</tbody></table></div>
</section>

<section class="sec" id="s-az">
  <p class="st">Piano di Recupero &amp; Azioni Correttive</p>
  <p class="ss">Diagnosi automatica &middot; Priorit&agrave; &middot; Piano settimanale</p>
  <div class="alerts">{alerts}</div>
  <div class="g2">
    <div class="tw"><div class="twh"><h3>&#x1F6A6; Delta vs Obiettivo</h3><span>Verde&ge;4 &middot; Arancione&ge;2 &middot; Rosso&lt;2</span></div>
      <table><thead><tr><th>Family Banker</th><th>Pol.</th><th>Obj/m</th><th>Stato</th><th>Gap</th></tr></thead>
      <tbody>{delta_rows}</tbody></table></div>
    <div class="tw"><div class="twh"><h3>&#x1F4CB; Piano d'Azione Settimanale</h3><span>Questa settimana</span></div>
      <table><thead><tr><th>Priorit&agrave;</th><th>Azione</th><th>Chi</th><th>Scadenza</th></tr></thead>
      <tbody>{piano_rows}</tbody></table></div>
  </div>
</section>

<section class="sec" id="s-pf">
  <p class="st">&#x1F4C1; Portafoglio Clienti</p>
  <p class="ss">Tutte le polizze sottoscritte YTD &middot; Data sottoscrizione e data incasso</p>

  <!-- Summary cards -->
  <div class="g4" style="margin-bottom:8px">
    <div class="card green">
      <div class="cico">&#x2705;</div>
      <p class="cl">Processate</p>
      <p class="cv">{pf_proc}</p>
      <p class="csub">Incasso completato</p>
    </div>
    <div class="card amb">
      <div class="cico">&#x23F3;</div>
      <p class="cl">In Lavorazione</p>
      <p class="cv">{pf_lav}</p>
      <p class="csub">In attesa di assunzione</p>
    </div>
    <div class="card red">
      <div class="cico">&#x274C;</div>
      <p class="cl">Annullate</p>
      <p class="cv">{pf_ann}</p>
      <p class="csub">Non assunte dalla compagnia</p>
    </div>
  </div>
  <div class="g4" style="margin-bottom:16px">
    <div class="card">
      <div class="cico">&#x1F4C5;</div>
      <p class="cl">Polizze a Pagamento Mensile</p>
      <p class="cv">{pf_mensile}</p>
      <p class="csub">Frazionamento mensile</p>
    </div>
    <div class="card">
      <div class="cico">&#x1F4B0;</div>
      <p class="cl">Polizze a Pagamento Annuale</p>
      <p class="cv">{pf_annuale}</p>
      <p class="csub">Versamento unico</p>
    </div>
    <div class="card gold">
      <div class="cico">&#x1F4CA;</div>
      <p class="cl">Premio Medio per Polizza</p>
      <p class="cv">{pf_med}</p>
      <p class="csub">Solo polizze processate</p>
    </div>
  </div>

  <!-- Filter bar -->
  <div style="display:flex;gap:8px;margin-bottom:12px;flex-wrap:wrap;align-items:center">
    <input type="text" id="pf-search" placeholder="&#x1F50D; Cerca FB o cliente..." oninput="filterPF()"
      style="padding:7px 12px;border:1px solid var(--brd);border-radius:6px;font-family:'DM Sans',sans-serif;font-size:.8rem;background:var(--cream);color:var(--navy);min-width:220px">
    <select id="pf-stato" onchange="filterPF()"
      style="padding:7px 12px;border:1px solid var(--brd);border-radius:6px;font-family:'DM Sans',sans-serif;font-size:.8rem;background:var(--cream);color:var(--navy)">
      <option value="">Tutti gli stati</option>
      <option value="Processata">&#x2713; Processata</option>
      <option value="In lavorazione">&#x23F3; In lavorazione</option>
      <option value="Annullata">&#x274C; Annullata</option>
    </select>
    <select id="pf-fb" onchange="filterPF()"
      style="padding:7px 12px;border:1px solid var(--brd);border-radius:6px;font-family:'DM Sans',sans-serif;font-size:.8rem;background:var(--cream);color:var(--navy)">
      <option value="">Tutti i FB</option>
    </select>
    <span id="pf-count" style="font-size:.75rem;color:var(--mut);margin-left:4px"></span>
  </div>

  <div style="overflow-x:auto">
    <table id="pf-table">
      <thead><tr>
        <th>Family Banker</th>
        <th>Cliente</th>
        <th>Prodotto</th>
        <th>Frazionamento</th>
        <th>Premio<br>alla Firma</th>
        <th>Data<br>Sottoscrizione</th>
        <th>Data<br>Incasso</th>
        <th>Stato</th>
      </tr></thead>
      <tbody id="pf-tbody">{pf_rows}</tbody>
    </table>
  </div>
</section>

<script>
(function(){{
  // Populate FB filter
  var rows = document.querySelectorAll('#pf-tbody tr');
  var fbs = new Set();
  rows.forEach(function(r){{ var td = r.querySelector('td'); if(td) fbs.add(td.textContent.trim()); }});
  var sel = document.getElementById('pf-fb');
  if(sel) fbs.forEach(function(fb){{ var o=document.createElement('option'); o.value=fb; o.textContent=fb; sel.appendChild(o); }});
  updatePFcount();
}})();

function filterPF(){{
  var q     = (document.getElementById('pf-search').value||'').toLowerCase();
  var stato = (document.getElementById('pf-stato').value||'').toLowerCase();
  var fb    = (document.getElementById('pf-fb').value||'').toLowerCase();
  var rows  = document.querySelectorAll('#pf-tbody tr');
  rows.forEach(function(r){{
    var txt = r.textContent.toLowerCase();
    var show = (!q || txt.includes(q)) &&
               (!stato || txt.includes(stato)) &&
               (!fb || r.querySelector('td').textContent.trim().toLowerCase() === fb);
    r.style.display = show ? '' : 'none';
  }});
  updatePFcount();
}}
function updatePFcount(){{
  var vis = document.querySelectorAll('#pf-tbody tr:not([style*="none"])').length;
  var tot = document.querySelectorAll('#pf-tbody tr').length;
  var el  = document.getElementById('pf-count');
  if(el) el.textContent = vis + ' di ' + tot + ' polizze';
}}
</script>



<section class="sec" id="s-obj">
  <p class="st">&#x1F4C8; Obiettivi Settimanali 2026</p>
  <p class="ss">Avanzamento incasso settimanale vs target &middot; Stima proporzionale al premio annuo</p>
  <div class="g2">
    <div class="tw">
      <div class="twh"><h3>&#x1F3AF; Budget Personale &mdash; Target {fe_target_300}</h3><span>Settimana per settimana</span></div>
      <table><thead><tr><th>Sett.</th><th>Pol.</th><th>Inc. Sett.</th><th>% Sett.</th><th>Cumulato</th><th>% Cum.</th></tr></thead>
      <tbody>{obj_rows_300}</tbody>
      <tfoot><tr style="background:rgba(11,30,61,.06);font-weight:600">
        <td>TOTALE</td><td class="num">{tot_pol_obj}</td>
        <td class="num">{fe_tot_inc_300}</td><td></td>
        <td class="num">{fe_tot_inc_300}</td>
        <td><span class="num" style="color:{col_fin_300};font-weight:700">{pct_fin_300}%</span></td>
      </tr></tfoot>
    </table></div>
    <div class="tw">
      <div class="twh"><h3>&#x1F465; Gruppo Onorato &mdash; Target {fe_target_180}</h3><span>Settimana per settimana</span></div>
      <table><thead><tr><th>Sett.</th><th>Pol.</th><th>Inc. Sett.</th><th>% Sett.</th><th>Cumulato</th><th>% Cum.</th></tr></thead>
      <tbody>{obj_rows_180}</tbody>
      <tfoot><tr style="background:rgba(11,30,61,.06);font-weight:600">
        <td>TOTALE</td><td class="num">{tot_pol_obj}</td>
        <td class="num">{fe_tot_inc_180}</td><td></td>
        <td class="num">{fe_tot_inc_180}</td>
        <td><span class="num" style="color:{col_fin_180};font-weight:700">{pct_fin_180}%</span></td>
      </tr></tfoot>
    </table></div>
  </div>
</section>

<section class="sec" id="s-ch">
  <p class="st">&#x1F3C6; Challenge</p>
  <p class="ss">Configura partecipanti, periodo e obiettivi &middot; Classifica da Excel</p>

  <div style="display:flex;gap:8px;margin-bottom:16px;border-bottom:2px solid var(--brd)">
    <button class="ch-tab on" onclick="showChTab('standard')" style="padding:8px 20px;border:none;border-bottom:3px solid var(--gold);background:transparent;font-family:'DM Sans',sans-serif;font-size:.82rem;font-weight:600;color:var(--navy);cursor:pointer;margin-bottom:-2px">&#x1F3C6; Challenge Corrente</button>
    <button class="ch-tab" onclick="showChTab('londra')" style="padding:8px 20px;border:none;border-bottom:3px solid transparent;background:transparent;font-family:'DM Sans',sans-serif;font-size:.82rem;font-weight:500;color:var(--mut);cursor:pointer;margin-bottom:-2px">&#x1F1EC;&#x1F1E7; Challenge Londra</button>
  </div>

  <div id="ch-standard">
  <!-- CONFIGURATORE CHALLENGE -->
  <div class="tw" style="margin-bottom:16px">
    <div class="twh"><h3>&#x2699;&#xFE0F; Configura Nuova Challenge</h3><span>Seleziona parametri e aggiorna Excel per salvare</span></div>
    <div style="padding:20px 22px">
      <div class="g3" style="margin-bottom:16px">
        <div>
          <p style="font-size:.65rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:6px">Data Inizio</p>
          <input type="date" id="ch-inizio" value="2026-04-09"
            style="width:100%;padding:8px 10px;border:1px solid var(--brd);border-radius:6px;font-family:'DM Sans',sans-serif;font-size:.82rem;background:var(--cream);color:var(--navy)">
        </div>
        <div>
          <p style="font-size:.65rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:6px">Data Fine</p>
          <input type="date" id="ch-fine" value="2026-05-09"
            style="width:100%;padding:8px 10px;border:1px solid var(--brd);border-radius:6px;font-family:'DM Sans',sans-serif;font-size:.82rem;background:var(--cream);color:var(--navy)">
        </div>
        <div>
          <p style="font-size:.65rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:6px">Obiettivi per Vincere</p>
          <div style="display:flex;gap:8px">
            <div style="flex:1">
              <label style="font-size:.68rem;color:var(--mut)">Min. Polizze</label>
              <input type="number" id="ch-min-pol" value="4" min="0"
                style="width:100%;padding:6px 8px;border:1px solid var(--brd);border-radius:5px;font-size:.82rem;background:var(--cream);margin-top:3px">
            </div>
            <div style="flex:1">
              <label style="font-size:.68rem;color:var(--mut)">Min. Premio (€)</label>
              <input type="number" id="ch-min-pa" value="2500" min="0" step="100"
                style="width:100%;padding:6px 8px;border:1px solid var(--brd);border-radius:5px;font-size:.82rem;background:var(--cream);margin-top:3px">
            </div>
          </div>
        </div>
      </div>
      <div>
        <p style="font-size:.65rem;text-transform:uppercase;letter-spacing:.07em;color:var(--mut);margin-bottom:8px">Partecipanti — seleziona chi include nella challenge</p>
        <div id="ch-partecipanti" style="display:flex;flex-wrap:wrap;gap:6px">
          <label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Alzetta Stefano' onchange='updateChallenge()' style='accent-color:var(--gold)'>Alzetta Stefano</label><label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Battisti Amelia' onchange='updateChallenge()' style='accent-color:var(--gold)'>Battisti Amelia</label><label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Cangemi Francesco' onchange='updateChallenge()' style='accent-color:var(--gold)'>Cangemi Francesco</label><label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Cimaschi Matteo' onchange='updateChallenge()' style='accent-color:var(--gold)'>Cimaschi Matteo</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Cristofari Rita' onchange='updateChallenge()' style='accent-color:var(--gold)'>Cristofari Rita</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Curra Isabella' onchange='updateChallenge()' style='accent-color:var(--gold)'>Curra Isabella</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='De Marco Fabio' onchange='updateChallenge()' style='accent-color:var(--gold)'>De Marco Fabio</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Di Marcello Sabrina' onchange='updateChallenge()' style='accent-color:var(--gold)'>Di Marcello Sabrina</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Ercolino Giovanni' onchange='updateChallenge()' style='accent-color:var(--gold)'>Ercolino Giovanni</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Galieti Fabio' onchange='updateChallenge()' style='accent-color:var(--gold)'>Galieti Fabio</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Gallo Armando' onchange='updateChallenge()' style='accent-color:var(--gold)'>Gallo Armando</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Gallo Tiziano' onchange='updateChallenge()' style='accent-color:var(--gold)'>Gallo Tiziano</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Grifoni Cristiana' onchange='updateChallenge()' style='accent-color:var(--gold)'>Grifoni Cristiana</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Lolletti Benedetto' onchange='updateChallenge()' style='accent-color:var(--gold)'>Lolletti Benedetto</label><label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Maola Daniele' onchange='updateChallenge()' style='accent-color:var(--gold)'>Maola Daniele</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Mazzetta Alessandra' onchange='updateChallenge()' style='accent-color:var(--gold)'>Mazzetta Alessandra</label><label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Onorato Beniamino' onchange='updateChallenge()' style='accent-color:var(--gold)'>Onorato Beniamino</label><label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Panfili Stefano' onchange='updateChallenge()' style='accent-color:var(--gold)'>Panfili Stefano</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Pozzana Patrizia' onchange='updateChallenge()' style='accent-color:var(--gold)'>Pozzana Patrizia</label><label style='display:inline-flex;align-items:center;gap:5px;background:rgba(200,169,81,.15);border:1px solid var(--gold);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox' checked value='Rizza Andrea' onchange='updateChallenge()' style='accent-color:var(--gold)'>Rizza Andrea</label><label style='display:inline-flex;align-items:center;gap:5px;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:4px 10px;cursor:pointer;font-size:.75rem;user-select:none'><input type='checkbox'  value='Santachiara Manuel' onchange='updateChallenge()' style='accent-color:var(--gold)'>Santachiara Manuel</label>
        </div>
      </div>
      <div style="margin-top:16px;padding-top:16px;border-top:1px solid var(--brd);display:flex;gap:10px;align-items:center">
        <button onclick="resetChallenge()"
          style="padding:8px 20px;background:var(--cream);border:1px solid var(--brd);border-radius:6px;cursor:pointer;font-family:'DM Sans',sans-serif;font-size:.8rem;color:var(--navy)">
          &#x21BA; Reset selezione Excel
        </button>
        <p style="font-size:.72rem;color:var(--mut)">&#x2139;&#xFE0F; La classifica viene sempre letta dall'Excel. Questo configuratore mostra un'anteprima dei partecipanti selezionati.</p>
      </div>
    </div>
  </div>

  <div class="g2" style="align-items:start">
    <div class="tw">
      <div class="twh"><h3>&#x1F4CA; Classifica Challenge</h3><span id="ch-periodo-lbl">09/04/2026 - 09/05/2026</span></div>
      <div style="padding:10px 18px;background:var(--cream);border-bottom:1px solid var(--brd);display:flex;gap:16px;flex-wrap:wrap">
        <span class="tag bb">Min. Polizze: 4</span>
        <span class="tag bn">Min. Premio: € 2.500</span>
      </div>
      <table><thead><tr><th>Pos.</th><th>Family Banker</th><th>Polizze</th><th>Premio Annuo</th><th>Stato</th></tr></thead>
      <tbody id="ch-classifica"><tr style='background:linear-gradient(90deg,rgba(200,169,81,.1),transparent)'><td>&#x1F947;</td><td><strong>Grifoni Cristiana</strong></td><td><span class="tag tg">&#x2713; 5 pol.</span></td><td><span class="tag tg">&#x2713; € 3.091</span></td><td style='font-size:.8rem'>&#x1F3C6; Vincitore</td></tr><tr style=''><td>&#x1F948;</td><td><strong>Pozzana Patrizia</strong></td><td><span class="tag tg">&#x2713; 7 pol.</span></td><td><span class="tag ta">€ 2.191</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>&#x1F949;</td><td><strong>Rizza Andrea</strong></td><td><span class="tag ta">2 pol.</span></td><td><span class="tag ta">€ 1.698</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>4&#xB0;</td><td><strong>Galieti Fabio</strong></td><td><span class="tag ta">3 pol.</span></td><td><span class="tag ta">€ 1.667</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>5&#xB0;</td><td><strong>Curra Isabella</strong></td><td><span class="tag ta">1 pol.</span></td><td><span class="tag ta">€ 1.152</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>6&#xB0;</td><td><strong>Mazzetta Alessandra</strong></td><td><span class="tag ta">2 pol.</span></td><td><span class="tag ta">€ 964</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>7&#xB0;</td><td><strong>Cristofari Rita</strong></td><td><span class="tag ta">3 pol.</span></td><td><span class="tag ta">€ 740</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>8&#xB0;</td><td><strong>Ercolino Giovanni</strong></td><td><span class="tag ta">2 pol.</span></td><td><span class="tag ta">€ 673</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>9&#xB0;</td><td><strong>Gallo Armando</strong></td><td><span class="tag ta">1 pol.</span></td><td><span class="tag ta">€ 515</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>10&#xB0;</td><td><strong>Lolletti Benedetto</strong></td><td><span class="tag ta">2 pol.</span></td><td><span class="tag ta">€ 290</span></td><td style='font-size:.8rem'>&#x23F3; In corsa</td></tr><tr style=''><td>11&#xB0;</td><td><strong>De Marco Fabio</strong></td><td><span class="tag tr2">0 pol.</span></td><td><span class="tag tr2">€ 0</span></td><td style='font-size:.8rem'>&#x274C; Nessuna pol.</td></tr><tr style=''><td>12&#xB0;</td><td><strong>Di Marcello Sabrina</strong></td><td><span class="tag tr2">0 pol.</span></td><td><span class="tag tr2">€ 0</span></td><td style='font-size:.8rem'>&#x274C; Nessuna pol.</td></tr><tr style=''><td>13&#xB0;</td><td><strong>Gallo Tiziano</strong></td><td><span class="tag tr2">0 pol.</span></td><td><span class="tag tr2">€ 0</span></td><td style='font-size:.8rem'>&#x274C; Nessuna pol.</td></tr></tbody></table>
    </div>
    <div>
      <div class="tw">
        <div class="twh"><h3>&#x1F465; Partecipanti Selezionati</h3><span id="ch-n-part">13 selezionati</span></div>
        <div id="ch-pills-box" style="padding:14px 18px;line-height:2"><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>De Marco Fabio</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Curra Isabella</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Lolletti Benedetto</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Grifoni Cristiana</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Gallo Armando</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Mazzetta Alessandra</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Ercolino Giovanni</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Rizza Andrea</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Galieti Fabio</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Di Marcello Sabrina</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Gallo Tiziano</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Cristofari Rita</span><span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>Pozzana Patrizia</span></div>
      </div>
      <div class="card gold" style="margin-top:0">
        <div class="cico">&#x1F3C6;</div>
        <p class="cl">Obiettivi per vincere</p>
        <div style="margin-top:8px" id="ch-obj-box"><p style='font-size:.8rem;margin-bottom:6px'>&#x1F4CB; <strong>Polizze minime:</strong> 4</p><p style='font-size:.8rem'>&#x1F4B6; <strong>Premio annuo minimo:</strong> € 2.500</p></div>
      </div>
    </div>
  </div>
  </div><!-- end ch-standard -->

  <div id="ch-londra" style="display:none">
    <div style="background:linear-gradient(135deg,var(--navy),var(--n3));border-radius:12px;padding:20px 24px;margin-bottom:16px;display:flex;align-items:center;gap:16px;box-shadow:0 4px 20px rgba(11,30,61,.18)">
      <div style="font-size:2.5rem;line-height:1">&#x1F1EC;&#x1F1E7;</div>
      <div>
        <p style="font-family:'Playfair Display',serif;color:var(--g2);font-size:1.2rem;font-weight:700;margin-bottom:3px">Challenge Londra 2026</p>
        <p style="color:rgba(255,255,255,.55);font-size:.75rem">1 Maggio &mdash; 30 Giugno 2026 &middot; Gruppo Onorato</p>
        <p style="color:rgba(255,255,255,.38);font-size:.67rem;margin-top:4px;font-style:italic">Solo incassato di nuove polizze sottoscritte nel periodo: non contano le rate di polizze precedenti</p>
      </div>
      <div style="margin-left:auto;text-align:right;flex-shrink:0">
        <p style="font-family:'Outfit',sans-serif;color:var(--gold);font-size:1.4rem;font-weight:700;line-height:1">&#x20AC; 5.000</p>
        <p style="color:rgba(255,255,255,.45);font-size:.63rem;margin-top:2px">Min. incassato</p>
        <p style="font-family:'Outfit',sans-serif;color:var(--g2);font-size:1rem;font-weight:600;margin-top:6px;line-height:1">4 polizze</p>
        <p style="color:rgba(255,255,255,.45);font-size:.63rem;margin-top:2px">Min. polizze</p>
      </div>
    </div>
    <div class="g2" style="align-items:start">
      <div class="tw">
        <div class="twh"><h3>&#x1F4CA; Classifica Challenge Londra</h3><span>01/05/2026 &ndash; 30/06/2026</span></div>
        <div style="padding:10px 18px;background:var(--cream);border-bottom:1px solid var(--brd);display:flex;gap:16px;flex-wrap:wrap;align-items:center">
          <span class="tag bb">Min. Polizze: 4</span>
          <span class="tag bn">Min. Incassato: &#x20AC; 5.000</span>
          {lon_stato_bar}
        </div>
        <table style="white-space:nowrap;font-size:.75rem"><thead><tr><th>Pos.</th><th>Family Banker</th><th>N&#xB0; Pol.</th><th>Incassato</th><th>Manca a € 5.000</th><th style="min-width:80px">Stato</th></tr></thead>
        <tbody>{lon_rows}</tbody></table>
      </div>
      <div>
        <div class="tw">
          <div class="twh"><h3>&#x1F465; Partecipanti</h3><span>Gruppo Onorato ({lon_n_part})</span></div>
          <div style="padding:14px 18px;line-height:2">{lon_pills}</div>
        </div>
        <div class="card gold" style="margin-top:12px">
          <div class="cico">&#x1F1EC;&#x1F1E7;</div>
          <p class="cl">Obiettivi doppi per vincere</p>
          <div style="margin-top:8px">
            <p style="font-size:.8rem;margin-bottom:6px">&#x1F4CB; <strong>Min. polizze:</strong> 4 nuove polizze nel periodo</p>
            <p style="font-size:.8rem;margin-bottom:6px">&#x1F4B6; <strong>Min. incassato:</strong> &#x20AC; 5.000 (rate &#x2713; di mag+giu)</p>
            <p style="font-size:.75rem;color:var(--mut);margin-top:8px;line-height:1.5">La classifica usa l&#x2019;incassato reale delle polizze nuove, non il premio annuo.</p>
          </div>
        </div>
      </div>
    </div>
  </div><!-- end ch-londra -->

</section>

</div>
<footer style="background:var(--navy);color:rgba(255,255,255,.3);text-align:center;padding:14px;font-size:.65rem;letter-spacing:.05em;border-top:1px solid var(--gold)">
  <span style="color:var(--gold)">FAMILY PROTECTION SPECIALIST</span> &nbsp;&middot;&nbsp; Dashboard KPI 2026 &nbsp;&middot;&nbsp; Aggiornato il {oggi} &nbsp;&middot;&nbsp; Uso interno riservato
</footer>

<script>
var COLLDATA = {coll_json};
function showSec(id) {{
  document.querySelectorAll('.sec').forEach(function(s){{s.classList.remove('on');}});
  document.querySelectorAll('.nb').forEach(function(b){{b.classList.remove('on');}});
  document.getElementById('s-'+id).classList.add('on');
  var idx={{ov:0,fb:1,tr:2,obj:3,ch:4,az:5}}[id];
  if(idx!==undefined) document.querySelectorAll('.nb')[idx].classList.add('on');
}}
function showColl(name) {{
  var body=document.getElementById('collb');
  if(!name){{body.innerHTML='<div class="collph">Seleziona un Family Banker per la scheda completa del colloquio</div>';return;}}
  body.innerHTML=COLLDATA[name]||'<div class="collph">Dati non disponibili</div>';
}}

// ─── CHALLENGE INTERATTIVA ───────────────────────────────────────────
var CH_DATI = {ch_dati_json};

function fmtEur(v) {{
  var s=String(Math.round(v||0));
  var r=''; var c=0;
  for(var i=s.length-1;i>=0;i--){{r=s[i]+(c>0&&c%3===0?'.':'')+r;c++;}}
  return r;
}}

function updateChallenge() {{
  // Partecipanti selezionati
  var sel = [];
  document.querySelectorAll('#ch-partecipanti input[type=checkbox]:checked').forEach(function(cb){{sel.push(cb.value);}});
  
  // Aggiorna pills
  var pillsBox = document.getElementById('ch-pills-box');
  if(pillsBox) pillsBox.innerHTML = sel.map(function(nm){{
    return "<span style='display:inline-block;background:var(--cream);border:1px solid var(--brd);border-radius:20px;padding:3px 10px;font-size:.72rem;margin:2px'>"+nm+"</span>";
  }}).join('');
  document.getElementById('ch-n-part').textContent = sel.length + ' selezionati';

  // Obiettivi
  var minPol = parseInt(document.getElementById('ch-min-pol').value)||0;
  var minPa  = parseFloat(document.getElementById('ch-min-pa').value)||0;

  // Aggiorna obj box
  var objBox = document.getElementById('ch-obj-box');
  var objHtml = '';
  if(minPol>0) objHtml += "<p style='font-size:.8rem;margin-bottom:6px'>&#x1F4CB; <strong>Polizze minime:</strong> "+minPol+"</p>";
  if(minPa>0)  objHtml += "<p style='font-size:.8rem'>&#x1F4B6; <strong>Premio annuo minimo:</strong> &#x20AC;&#x00A0;"+Math.round(minPa).toLocaleString('it-IT')+"</p>";
  if(!objHtml) objHtml = "<p style='font-size:.8rem;color:#64748B'>Nessun obiettivo impostato</p>";
  if(objBox) objBox.innerHTML = objHtml;

  // Periodo
  var ini  = document.getElementById('ch-inizio').value;
  var fine = document.getElementById('ch-fine').value;
  var periodoLbl = ini && fine ? ini.split('-').reverse().join('/') + ' - ' + fine.split('-').reverse().join('/') : '';
  var pl = document.getElementById('ch-periodo-lbl');
  if(pl && periodoLbl) pl.textContent = periodoLbl;

  // Filtra e ordina classifica solo per i selezionati
  var filtrati = CH_DATI.filter(function(r){{return sel.indexOf(r.fb)>=0;}});
  filtrati.sort(function(a,b){{return b.pa-a.pa || b.pol-a.pol;}});
  
  var medals = ['&#x1F947;','&#x1F948;','&#x1F949;'];
  var tbody = document.getElementById('ch-classifica');
  if(!tbody) return;
  tbody.innerHTML = filtrati.map(function(row, i){{
    var vp = minPol>0 && row.pol>=minPol;
    var va = minPa>0 && row.pa>=minPa;
    var vince = vp && (minPa===0 || va);
    var bg = vince ? "background:linear-gradient(90deg,rgba(200,169,81,.1),transparent)" : "";
    var bpol = vp ? "<span class='tag tg'>&#x2713; "+row.pol+" pol.</span>" : (row.pol>0 ? "<span class='tag ta'>"+row.pol+" pol.</span>" : "<span class='tag tr2'>0 pol.</span>");
    var bpa  = va  ? "<span class='tag tg'>&#x2713; &#x20AC;&#x00A0;"+fmtEur(Math.round(row.pa))+"</span>" : (row.pa>0 ? "<span class='tag ta'>&#x20AC;&#x00A0;"+fmtEur(Math.round(row.pa))+"</span>" : "<span class='tag tr2'>&#x20AC;&#x00A0;0</span>");
    var stato = vince ? "&#x1F3C6; Vincitore" : (row.pol>0 ? "&#x23F3; In corsa" : "&#x274C; Nessuna pol.");
    var med = i<3 ? medals[i] : (i+1)+"&#xB0;";
    return "<tr style='"+bg+"'><td>"+med+"</td><td><strong>"+row.fb+"</strong></td><td>"+bpol+"</td><td>"+bpa+"</td><td style='font-size:.8rem'>"+stato+"</td></tr>";
  }}).join('');

  // Stile checkbox
  document.querySelectorAll('#ch-partecipanti label').forEach(function(lbl){{
    var cb = lbl.querySelector('input');
    if(cb.checked) {{
      lbl.style.background='rgba(200,169,81,.15)';
      lbl.style.borderColor='var(--gold)';
    }} else {{
      lbl.style.background='var(--cream)';
      lbl.style.borderColor='var(--brd)';
    }}
  }});
}}

function resetChallenge() {{
  // Ripristina partecipanti originali da Excel
  var originali = {ch_originali_json};
  document.querySelectorAll('#ch-partecipanti input[type=checkbox]').forEach(function(cb){{
    cb.checked = originali.indexOf(cb.value) >= 0;
  }});
  document.getElementById('ch-min-pol').value = '{ch_min_pol}';
  document.getElementById('ch-min-pa').value = '{ch_min_pa_val_js}';
  updateChallenge();
}}

// Init challenge al caricamento
document.addEventListener('DOMContentLoaded', function(){{ updateChallenge(); }});

function showChTab(tab) {{
  var isLondra = tab === 'londra';
  document.getElementById('ch-standard').style.display = isLondra ? 'none' : 'block';
  document.getElementById('ch-londra').style.display   = isLondra ? 'block' : 'none';
  document.querySelectorAll('.ch-tab').forEach(function(btn, i) {{
    var active = (i===0 && !isLondra) || (i===1 && isLondra);
    btn.style.borderBottomColor = active ? 'var(--gold)' : 'transparent';
    btn.style.fontWeight = active ? '600' : '500';
    btn.style.color = active ? 'var(--navy)' : 'var(--mut)';
  }});
}}
</script>

<script>
document.addEventListener('DOMContentLoaded', function(){{
  var FBS = Object.keys(COLLDATA).sort();
  var selFbs = new Set();
  var cbDiv = document.getElementById('prod-fb-checkboxes');
  FBS.forEach(function(fb){{
    var lbl = document.createElement('label');
    lbl.style.cssText = 'display:flex;align-items:center;gap:7px;padding:5px 4px;cursor:pointer;font-size:.8rem;';
    var cb = document.createElement('input');
    cb.type='checkbox'; cb.value=fb;
    cb.addEventListener('change', function(){{ if(this.checked) selFbs.add(fb); else selFbs.delete(fb); updatePills(); }});
    lbl.appendChild(cb); lbl.appendChild(document.createTextNode(' '+fb));
    cbDiv.appendChild(lbl);
  }});
  document.getElementById('prod-dal').value = '2026-01-01';
  document.getElementById('prod-al').value = new Date().toISOString().slice(0,10);
  window.toggleProdFbDropdown = function(){{
    var d=document.getElementById('prod-fb-dropdown');
    d.style.display = d.style.display==='none'?'block':'none';
  }};
  document.addEventListener('click',function(e){{
    if(!e.target.closest('#prod-fb-pills')&&!e.target.closest('#prod-fb-dropdown')){{
      var d=document.getElementById('prod-fb-dropdown'); if(d) d.style.display='none';
    }}
  }});
  window.prodSelectAll=function(){{ selFbs=new Set(FBS); cbDiv.querySelectorAll('input').forEach(function(c){{c.checked=true;}}); updatePills(); }};
  window.prodSelectNone=function(){{ selFbs=new Set(); cbDiv.querySelectorAll('input').forEach(function(c){{c.checked=false;}}); updatePills(); }};
  function updatePills(){{
    var box=document.getElementById('prod-fb-pills');
    var ph=document.getElementById('prod-fb-placeholder');
    box.querySelectorAll('.ppill').forEach(function(p){{p.remove();}});
    ph.style.display = selFbs.size===0 ? 'inline' : 'none';
    selFbs.forEach(function(fb){{
      var sp=document.createElement('span'); sp.className='ppill';
      sp.style.cssText='background:var(--navy);color:#fff;border-radius:14px;padding:3px 10px;font-size:.72rem;display:inline-flex;align-items:center;gap:5px;';
      var ini=fb.split(' ').map(function(w){{return w[0];}}).join('').slice(0,2).toUpperCase();
      sp.appendChild(document.createTextNode(ini+' '));
      var x=document.createElement('span'); x.innerHTML='&times;'; x.style.cssText='cursor:pointer;opacity:.7;';
      x.onclick=function(e){{ e.stopPropagation(); selFbs.delete(fb); var c=cbDiv.querySelector('input[value="'+fb+'"]'); if(c)c.checked=false; updatePills(); }};
      sp.appendChild(x); box.appendChild(sp);
    }});
  }}
  function fmtE(n){{ var s=String(Math.round(n||0)),r='',c=0; for(var i=s.length-1;i>=0;i--){{r=s[i]+(c>0&&c%3===0?'.':'')+r;c++;}} return '\u20ac\u00a0'+r; }}
  function pf2(s){{ return parseFloat(String(s||0).replace(/\./g,'').replace(',','.'))||0; }}
  function parseDate(s){{ if(!s)return null; var p=s.split('/'); return p.length===3?new Date(+p[2],+p[1]-1,+p[0]):null; }}
  function getPols(fb){{
    var h=COLLDATA[fb]||'', res=[];
    var re=/<tr><td>([\d\/]*)<\/td><td><strong>[^<]*<\/strong><\/td><td>[^<]*<\/td><td class="num">[^<]*<\/td><td class="num">[^\u20ac]*[\u20ac\u00a0\s]*([\d\.,]*)<\/td><td>[^<]*<\/td><td><span class="tag ([^"]*)"/g, m;
    while((m=re.exec(h))!==null) res.push({{date:m[1],pa:pf2(m[2]),lav:m[3].indexOf('tb')>=0}});
    return res;
  }}
  function getAppts(fb){{ var h=COLLDATA[fb]||'', m=h.match(/<tr><td>Appt<\/td>(.*?)<\/tr>/), ns; if(!m)return 0; ns=m[1].match(/>([\d]+)</g); return ns?ns.reduce(function(s,n){{return s+parseInt(n.replace(/[><]/g,''));}},0):0; }}
  function getCb(fb){{ var h=COLLDATA[fb]||'', m=h.match(/Callback<\/span><span class="csv"[^>]*>([\d]+)/); return m?+m[1]:0; }}
  function getInc(fb){{ var h=COLLDATA[fb]||'', m=h.match(/Premi Incassati<\/p><p class="ckv num"[^>]*>[^\u20ac]*[\u20ac\u00a0\s]*([\d\.\,]+)/); return m?pf2(m[1]):0; }}
  window.runProdAnalysis=function(){{
    if(!selFbs.size){{alert('Seleziona almeno un Family Banker');return;}}
    var dal=document.getElementById('prod-dal').value, al=document.getElementById('prod-al').value;
    if(!dal||!al){{alert('Seleziona il periodo');return;}}
    var dF=new Date(dal), dT=new Date(al); dT.setHours(23,59,59);
    var ms=[]; document.querySelectorAll('#prod-metrics input:checked').forEach(function(c){{ms.push(c.value);}});
    var rows=[],tpa=0,tinc=0,tpol=0,tappt=0,tcb=0,tlav=0;
    selFbs.forEach(function(fb){{
      var pols=getPols(fb).filter(function(p){{var d=parseDate(p.date);return d&&d>=dF&&d<=dT;}});
      var pa=pols.reduce(function(s,p){{return s+p.pa;}},0), pol=pols.length, lav=pols.filter(function(p){{return p.lav;}}).length;
      var inc=getInc(fb), appt=getAppts(fb), cb=getCb(fb);
      tpa+=pa;tinc+=inc;tpol+=pol;tappt+=appt;tcb+=cb;tlav+=lav;
      rows.push({{fb:fb,pa:pa,inc:inc,pol:pol,appt:appt,cb:cb,lav:lav,conv:appt>0?Math.round(pol/appt*100):0,med:pol>0?Math.round(pa/pol):0}});
    }});
    rows.sort(function(a,b){{return b.pa-a.pa;}});
    var kd=document.getElementById('prod-kpi-cards'); kd.innerHTML='';
    function kc(l,v,c){{var d=document.createElement('div');d.style.cssText='background:var(--cream);border-radius:9px;padding:12px 16px;border-left:3px solid '+c+';min-width:120px;flex:1';d.innerHTML='<p style="font-size:.62rem;text-transform:uppercase;color:var(--mut);margin-bottom:4px">'+l+'</p><p style="font-family:Outfit,sans-serif;font-size:1.2rem;font-weight:700;color:var(--navy)">'+v+'</p>';kd.appendChild(d);}}
    if(ms.includes('pa'))   kc('Premio Annuo',fmtE(tpa),'#2E8B5F');
    if(ms.includes('pol'))  kc('N\u00b0 Polizze',tpol,'#2E75B6');
    if(ms.includes('appt')) kc('N\u00b0 Appt YTD',tappt,'#D97706');
    if(ms.includes('inc'))  kc('Incassato YTD',fmtE(tinc),'#8B5CF6');
    if(ms.includes('cb'))   kc('Callback',tcb,'#C0392B');
    if(ms.includes('lav'))  kc('In Lavorazione',tlav,'#F59E0B');
    var cols=[
      {{k:'fb',  l:'Family Banker',f:function(v){{return '<strong>'+v+'</strong>';}}}},
      {{k:'pol', l:'N\u00b0 Pol.',f:function(v){{return '<span class="num">'+v+'</span>';}}}},
      {{k:'pa',  l:'Premio Annuo',f:function(v){{return '<span class="num">'+fmtE(v)+'</span>';}}}},
      {{k:'inc', l:'Incassato',   f:function(v){{return '<span class="num">'+fmtE(v)+'</span>';}}}},
      {{k:'appt',l:'N\u00b0 Appt',f:function(v){{return '<span class="num">'+v+'</span>';}}}},
      {{k:'cb',  l:'Callback',    f:function(v){{return '<span class="num">'+v+'</span>';}}}},
      {{k:'lav', l:'In Lav.',     f:function(v){{return '<span class="num">'+v+'</span>';}}}},
      {{k:'conv',l:'% Conv.',     f:function(v){{return '<span class="num">'+v+'%</span>';}}}},
      {{k:'med', l:'Premio Medio',f:function(v){{return '<span class="num">'+fmtE(v)+'</span>';}}}}
    ];
    var ac=cols.filter(function(c){{return c.k==='fb'||ms.includes(c.k);}});
    document.getElementById('prod-thead').innerHTML='<tr>'+ac.map(function(c){{return '<th>'+c.l+'</th>';}}).join('')+'</tr>';
    var tot={{fb:'<strong>TOTALE</strong>',pa:tpa,inc:tinc,pol:tpol,appt:tappt,cb:tcb,lav:tlav,conv:tappt>0?Math.round(tpol/tappt*100):0,med:tpol>0?Math.round(tpa/tpol):0}};
    document.getElementById('prod-tbody').innerHTML=
      rows.map(function(r,i){{return '<tr style="'+(i%2?'background:var(--cream)':'')+'">'+ ac.map(function(c){{return '<td>'+c.f(r[c.k])+'</td>';}}).join('')+'</tr>';}}).join('')+
      '<tr style="border-top:2px solid var(--brd);font-weight:600">'+ac.map(function(c){{return '<td>'+c.f(tot[c.k])+'</td>';}}).join('')+'</tr>';
    document.getElementById('prod-note').textContent='Polizze dal '+dal.split('-').reverse().join('/')+' al '+al.split('-').reverse().join('/')+' \u00b7 Appt/Cb/Incassato sono YTD';
    document.getElementById('prod-results').style.display='block';
    document.getElementById('prod-empty').style.display='none';
    document.getElementById('prod-fb-dropdown').style.display='none';
  }};
}});
</script>
</body></html>"""

os.makedirs('docs', exist_ok=True)
with open('docs/index.html', 'w', encoding='utf-8') as f:
    f.write(HTML)

print(f"✅ Dashboard generata: docs/index.html ({len(HTML):,} caratteri)")
print(f"   Polizze: {G['polizze']}, Premi Incassati: {fe(G['premiIncassati'])}")
print(f"   Aggiornato il: {oggi}")

